import math
import pandas as pd
import re
from openpyxl import load_workbook
import random
import numpy as np
import time

COLUMN_NAMES = ("Примечание", "№ п/п", "Маркировка", "Вид провода", "№ провода", "Сечение", "Цвет",
                "Длина, мм (± 3мм)", "Уплотнитель 1", "Длина трубки, L (мм) 1", "Длина трубки, L (мм) 2",
                "Частичное снятие 1", "Частичное снятие 2", "Наконечник 1", "Аппликатор 1", "Уплотнитель 2",
                "Наконечник 2", "Аппликатор 2", "Армировка 1 (Трубка ПВХ, Тр. Терм., изоляторы)",
                "Армировка 2 (Трубка ПВХ, Тр. Терм., изоляторы)")

def is_empty(value):
    if value is None or value == '' or value == ' ':
        return True
    else:
        return False

class HarnessChartReader:
    CABEL = "Кабель"
    CABEL_COLUMN = "Примечание"
    TERMINAL_1_COL = "Наконечник 1"
    TERMINAL_2_COL = "Наконечник 2"
    SEAL_1_COL = "Уплотнитель 1"
    SEAL_2_COL = "Уплотнитель 2"
    WIRE_CUT_1_COL = "Частичное снятие 1"
    WIRE_CUT_2_COL = "Частичное снятие 2"
    ARMIROVKA_1_COL = "Армировка 1 (Трубка ПВХ, Тр. Терм., изоляторы)"
    ARMIROVKA_2_COL = "Армировка 2 (Трубка ПВХ, Тр. Терм., изоляторы)"
    TERMINAL_1_COL_NUM = 13
    TERMINAL_2_COL_NUM = 19
    ARMIROVKA_1_COL_NUM = 9
    ARMIROVKA_2_COL_NUM = 15

    numeric_columns = [
        "Сечение",
        "Частичное снятие 1",
        "Частичное снятие 2",
    ]

    __file = 0
    __workbook_file = None
    __worksheet_file = None
    __dataframe_file = None

    def __init__(self, file=None):
        self.file = None

    def get_dataframe(self):
        return self.__dataframe_file

    def get_workbook(self):
        return self.workbook_file

    def load_file(self, file):
        if file is not None:
            self.file = file
            self.workbook_file = load_workbook(file)
            self.worksheet_file = self.workbook_file.active

    def __save_file(self):
        self.workbook_file.save(self.file)

    #TODO: check func, there is suspect it is not correct
    def __delete_gray_xlsx(self):
        ws = self.worksheet_file

        for row in ws.iter_rows():
            for cell in row:
                white_tint = 0.0

                color = cell.fill.start_color.index
                if cell.fill.start_color.tint != white_tint or (color != '00000000' and color != 'FFFFFFFF'and color != 0) or color == 8 or color == 22:
                    value = cell.value

                    if type(value) is float or type(value) is int:
                        if math.floor(value) > 16:
                            cell.value = None
                        else:
                            pass
                    elif type(value) is str:
                        cell.value = ''
                    else:
                        cell.value = None

    def __detect_cabels_xlsx(self):
        ws = self.worksheet_file

        delete_rows = False

        for merged_coordinates in ws.merged_cells.ranges:
            for merged_cell in ws[merged_coordinates.coord]:
                rows_list = []
                for cell in merged_cell:
                    temp_string = cell.value
                    if type(temp_string) is str and "кабель" in temp_string.lower():
                        rows_list = re.findall('\d+', merged_coordinates.coord)
                        delete_rows = True

                # write cabel in unmerged cells
                if delete_rows:
                    ws.unmerge_cells(merged_coordinates.coord)
                    for row in rows_list:
                        ws['A{}'.format(row)].value = "Кабель"
                    break

            delete_rows = False

    def __get_number_cabels_rows(self):
        rows = list()

        for merged_cell in self.worksheet_file.merged_cells.ranges:
            main_cell = self.worksheet_file[merged_cell.coord][0][0]
            if type(main_cell.value) is str and 'кабель' in main_cell.value.lower():
                last_cell = self.worksheet_file[merged_cell.coord][-1][-1]
                rows.append((main_cell.row, last_cell.row))

        return rows

    def __close_xlsx(self):
        self.workbook_file.save(self.file)
        self.worksheet_file = None

    def __get_first_row_to_read(self):
        for row in self.worksheet_file.iter_rows():
            for cell in row:
                value = cell.value
                if type(value) is str and value.lower() == "сечение":
                    return cell.row

    def __check_numeric_cols(self, numberic_cols):
        for col in numberic_cols:
            for idx, row in self.__dataframe_file.iterrows():
                if type(self.__dataframe_file.loc[idx, col]) is str:
                    return False

        return True

    def __delete_positions_without_length(self):
        # delete positions without length

        idx_to_drop = []

        LENGTH_WIRE_COLUMN = "Длина, мм (± 3мм)"
        for index, row in self.__dataframe_file.iterrows():
            for column in self.numeric_columns:
                if column in self.__dataframe_file:
                    self.__dataframe_file.loc[index, column] = float(self.__dataframe_file.loc[index, column])
            if LENGTH_WIRE_COLUMN in self.__dataframe_file:
                value = self.__dataframe_file.loc[index, LENGTH_WIRE_COLUMN]
                if value is not None and value != '' and value == value:
                    self.__dataframe_file.loc[index, LENGTH_WIRE_COLUMN] = int(
                        self.__dataframe_file.loc[index, LENGTH_WIRE_COLUMN])
                else:
                    idx_to_drop.append(index)

        # delete positions without length

        self.__dataframe_file.drop(index=idx_to_drop, inplace=True)

        # recover indecies
        self.__dataframe_file.index = pd.Index(range(self.__dataframe_file.shape[0]))

    def __get_paired_terminals_cabels_armirovka(self, start_row):
        rows_terminals = list()
        rows_cabels = list()
        rows_armirovka = list()

        for merged_cell in self.worksheet_file.merged_cells.ranges:
            main_cell = self.worksheet_file[merged_cell.coord][0][0]
            if main_cell.row > start_row:
                if main_cell.column == self.TERMINAL_1_COL_NUM or main_cell.column == self.TERMINAL_2_COL_NUM:
                    last_cell = self.worksheet_file[merged_cell.coord][-1][-1]
                    terminal_num = 1 if main_cell.column == self.TERMINAL_1_COL_NUM else 2
                    rows_terminals.append((terminal_num, main_cell.row, last_cell.row))

                if main_cell.column == self.ARMIROVKA_1_COL_NUM or main_cell.column == self.ARMIROVKA_2_COL_NUM:
                    last_cell = self.worksheet_file[merged_cell.coord][-1][-1]
                    armirovka_num = 1 if main_cell.column == self.ARMIROVKA_1_COL_NUM else 2
                    value = main_cell.value
                    rows_armirovka.append((value, armirovka_num, main_cell.row, last_cell.row))
                if type(main_cell.value) is str and 'кабель' in main_cell.value.lower():
                    last_cell = self.worksheet_file[merged_cell.coord][-1][-1]
                    rows_cabels.append((main_cell.row, last_cell.row))

        return (rows_terminals, rows_cabels, rows_armirovka)

    def __process_file(self):
        """
        Delete or add permanent columns

        This func deletes:
        - Delete rows and cols if NaN
        - Check if all columns exist
        :return:
        """

        # delete if column not in columns_names
        columns_to_drop = list()
        for column_name, column_data in self.__dataframe_file.iteritems():
            if column_name not in COLUMN_NAMES:
                """if type(column_name) is str and 'Unnamed' in column_name:
                    column_name = int(column_name[-1:])"""
                columns_to_drop.append(column_name)

        self.__dataframe_file.drop(columns_to_drop, axis=1, inplace=True)


        if any(column not in self.__dataframe_file for column in COLUMN_NAMES):
            for column in COLUMN_NAMES:
                if column not in self.__dataframe_file:
                    self.__dataframe_file[column] = 0

        self.__dataframe_file.dropna(axis=0, how='all', inplace=True)

        # fulfill NaN with 0s for numeric columns
        for numeric_column in self.numeric_columns:
            if numeric_column in self.__dataframe_file:
                self.__dataframe_file.loc[:, numeric_column] = \
                    self.__dataframe_file.loc[:, numeric_column].fillna(value=0)

        # fulfill NaN with empty string for other columns

        self.__dataframe_file.fillna(value='', inplace=True)

        self.__delete_positions_without_length()

    def __str_wire_num(self, wire_num):
        try:
            return str(int(wire_num))
        except:
            return str(wire_num)

    def __read_xl(self, start_row=0):
        cnt = 0
        ws = self.worksheet_file

        """
        for row in ws.iter_rows():
            cnt += 1
            if cnt == start_row:
                break
            else:
                for cell in row:
                    if cell.value != None:
                        cell.value = None
        """


        temp_df = pd.DataFrame(ws.values)
        temp_df.drop(list(range(0, start_row - 1)), axis=0, inplace=True)

        temp_df.dropna(inplace=True, how='all', axis=0)
        temp_df.dropna(inplace=True, how='all', axis=1)
        temp_df.index = pd.Index(range(temp_df.shape[0]))

        temp_df.rename(columns=temp_df.loc[0, :], inplace=True)
        temp_df.drop(index=[0, 1], inplace=True)
        temp_df.index = pd.Index(range(temp_df.shape[0]))
        # temp_df['№ провода'].apply(lambda wire_num: self.__str_wire_num(wire_num))

        return temp_df

    def __fulfill_cabels(self, start_row, cabels_rows):
        rows_to_drop = list()
        for row_pair in cabels_rows:
            max_len_idx = row_pair[0]
            max_len = self.__dataframe_file.loc[row_pair[0] - start_row - 2, "Длина, мм (± 3мм)"]
            for row in range(row_pair[0] + 1, row_pair[1] + 1):
                row_len = self.__dataframe_file.loc[row - start_row - 2, "Длина, мм (± 3мм)"]
                if row_len > max_len:
                    rows_to_drop.append(max_len_idx)
                    max_len = row_len
                    max_len_idx = row - start_row - 2
                else:
                    rows_to_drop.append(row - start_row - 2)
        self.__dataframe_file.drop(rows_to_drop, axis=0, inplace=True)
        self.__dataframe_file.index = pd.Index(range(self.__dataframe_file.shape[0]))

        """
            cabel_name = self.__dataframe_file.loc[row_pair[0] - start_row - 2, 'Вид провода']
            for row in range(row_pair[0], row_pair[1] + 1):
                wire_number = self.__dataframe_file.loc[row - start_row - 2, '№ провода']
                self.__dataframe_file.loc[row - start_row - 2, 'Вид провода'] = str(cabel_name) + ' ' + str(wire_number)
        """

    def __delete_paired_terminals(self, start_row, terminals_rows):
        for paired_rows in terminals_rows:
            terminal_num = paired_rows[0]

            if terminal_num == 1:
                cols = [self.ARMIROVKA_1_COL, self.SEAL_1_COL, self.TERMINAL_1_COL]
            else:
                cols = [self.ARMIROVKA_2_COL, self.SEAL_2_COL, self.TERMINAL_2_COL]

            for row in range(paired_rows[1], paired_rows[2] + 1):
                self.__dataframe_file.loc[row - start_row - 2, cols] = '', '', ''

    def __unpair_armirovka(self, start_row, armirovka_rows):
        for paired_rows in armirovka_rows:
            armirovka = paired_rows[0]
            armirovka_num = paired_rows[1]
            armirovka_col = self.ARMIROVKA_1_COL if armirovka_num == 1 else self.ARMIROVKA_2_COL

            for row in range(paired_rows[2], paired_rows[3] + 1):
                self.__dataframe_file.loc[row - start_row - 2, armirovka_col] = armirovka

    def __process_paired_cells(self, start_row, terminal_rows, cabels_rows, armirovka_rows):
        self.__delete_paired_terminals(start_row, terminal_rows)
        self.__fulfill_cabels(start_row, cabels_rows)
        self.__unpair_armirovka(start_row, armirovka_rows)


    def read_file_chart(self):
        if self.file.name.endswith('xlsx'):

            # xlsx processing

            self.__delete_gray_xlsx()

            # cabels_rows = self.__get_number_cabels_rows()

            row_start = self.__get_first_row_to_read()

            paired_terminals, paired_cabels, paired_armirovka = self.__get_paired_terminals_cabels_armirovka(row_start)

            # dataframe processing

            self.__dataframe_file = self.__read_xl(start_row=row_start)
            self.__close_xlsx()
            self.__process_file()

            self.__process_paired_cells(row_start, paired_terminals, paired_cabels, paired_armirovka)

class ProcessDataframe:
    chart = None
    chart_copy = None
    new_chart = None
    HARNESS_NUMBER_COL = "harness"
    AMOUNT_COL = "amount"
    KOMAX_COL = 'komax'
    MARKING_COL = "marking"
    TERMINAL_1_COL = 'wire_terminal_1'
    TERMINAL_2_COL = 'wire_terminal_2'
    TIME_COL = 'time'
    SQUARE_COL = 'wire_square'
    WIRE_TYPE_COL = 'wire_type'
    KAPPA_COL = 'kappa'
    SEAL_1_COL = 'wire_seal_1'
    SEAL_2_COL = 'wire_seal_2'
    ARMIROVKA_1_COL = 'armirovka_1'
    ARMIROVKA_2_COL = 'armirovka_2'
    WIRE_CUT_1_COL = 'wire_cut_length_1'
    WIRE_CUT_2_COL = 'wire_cut_length_2'


    __cols_name = {
        "wire_terminal_1": "wire_terminal_2",
        "wire_seal_1": "wire_seal_2",
        "wire_cut_length_1": "wire_cut_length_2",
        "aplicator_1": "aplicator_2",
        "armirovka_1": "armirovka_2",
        "tube_len_1": "tube_len_2",
    }

    def __init__(self, dataframe):
        self.chart = dataframe
        self.chart_copy = dataframe
        self.new_chart = pd.DataFrame()

    def __swap_sides(self, rows, cols=None, copy=False):
        if cols is None:
            cols = self.__cols_name
        for row in rows:
            for key, item in cols.items():

                if copy:
                    self.chart_copy.loc[row, [key, item]] = self.chart_copy.loc[row, [item, key]].values
                else:
                    self.chart.loc[row, [key, item]] = self.chart.loc[row, [item, key]].values

    def delete_word_contain(self, *args):
        columns = list(self.chart)
        for idx, row in self.chart.iterrows():
            for arg in args:
                for column in columns:
                    value = self.chart.loc[idx, column]
                    type_cell = type(value)
                    if type_cell is str and type(arg) is str:
                        if str(arg) in value or arg in value:
                            self.chart.loc[idx, column] = ''
                    elif (type_cell is int or float) and (type(arg) is int or float) and value == arg:
                        self.chart.loc[idx, column] = 0

    def __swap_cols(self, col1, col2):
        filtered = []
        second_filtered = []
        rows, cols = self.chart.shape
        rows_to_swap = []

        """
        h3 = -1
        if pairing:
            try:
                h3 = get_index(df, "Наконечник 3")[0]
            except:
                h3 = -1
        """

        for index, row in self.chart.iterrows():
            """
            if pairing and h3 != -1:
                if df.iloc[row, h3] is not None:
                    filtered.append(df.iloc[row, col1])
            """
            if self.chart.loc[index, col1] not in filtered:
                main_word = self.chart.loc[index, col1]
                if not empty(main_word) and main_word not in second_filtered:
                    filtered.append(main_word)
                    for idx, row_content in self.chart.iterrows():
                        left_word = self.chart.loc[idx, col1]
                        right_word = self.chart.loc[idx, col2]

                        if right_word == main_word and right_word != left_word and left_word not in filtered and \
                                idx not in rows_to_swap:
                            if not empty(left_word) and left_word not in second_filtered:
                                second_filtered.append(left_word)
                            rows_to_swap.append(idx)
                        elif left_word in filtered and index not in rows_to_swap:
                            if not empty(right_word) and right_word not in second_filtered:
                                second_filtered.append(right_word)
                elif not empty(main_word) and main_word in second_filtered and index not in rows_to_swap:
                    right_word = self.chart.loc[index, col2]
                    if right_word not in second_filtered:
                        rows_to_swap.append(index)
            else:
                right_word = self.chart.loc[index, col2]
                if not empty(right_word) and right_word not in second_filtered:
                    second_filtered.append(right_word)

        self.__swap_sides(rows_to_swap)

        """
        rows_to_swap_second = []
        for index, row in self.chart.iterrows():
            main_word = self.chart.loc[index, col1]
            ax_word = self.chart.loc[index, col2]
            if main_word in second_filtered and ax_word not in second_filtered and index not in rows_to_swap:
                rows_to_swap_second.append(index)

        self.__swap_sides(rows_to_swap_second)"""

    def __swap_cols_gaps(self, col1, col2):
        rows_to_swap = []
        rows = self.chart.shape[0]

        main_value = None
        for idx, row in self.chart.iterrows():
            if idx != rows - 1:
                curr_sec_value = self.chart.loc[idx, col2]
                curr_first_value = self.chart.loc[idx, col1]
                next_first_value = self.chart.loc[idx + 1, col1]

                if next_first_value == '' or next_first_value == ' ' or next_first_value is None:
                    if main_value is None:
                        main_value = curr_first_value
                elif curr_first_value != '' and curr_first_value != ' ' and curr_first_value is not None:
                    main_value = None

                if main_value is not None:
                    if curr_sec_value == main_value:
                        rows_to_swap.append(idx)

        self.__swap_sides(rows_to_swap)

    def __swap_rows(self, row1, row2):
        b, c = self.chart.iloc[row1].copy(), self.chart.iloc[row2].copy()
        self.chart.iloc[row1], self.chart.iloc[row2] = c, b

    def __divide_in_groups(self):
        self.chart['groups'] = ''

        for index, row in self.chart.iterrows():
            self.chart.loc[index, 'groups'] = str(group_by(self.chart.loc[index, 'wire_square']))

    def __correct_marking(self):
        MARKING_COL = 'marking'

        self.chart.loc[:, MARKING_COL] = self.chart.loc[:, MARKING_COL].apply(lambda x: self.__correct_mark_row(x))

    def __correct_mark_row(self, mark):
        return "Белый" if type(mark) is not None and ("б" in mark or "Б" in mark) else "Чёрный"

    def __fulfill_terminals(self):
        start_time = time.time()
        MARKING_COL = 'marking'
        list_1 = []
        list_2 = []
        first = {}
        second = {}

        for idx, row in self.chart.iterrows():
            value_1 = self.chart.loc[idx, 'wire_terminal_1']
            group = self.chart.loc[idx, 'groups']
            value_2 = self.chart.loc[idx, 'wire_terminal_2']
            marking = self.chart.loc[idx, MARKING_COL]

            if empty(value_1):
                for idx_local, row_local in self.chart.iterrows():
                    same_group = group == self.chart.loc[idx_local, 'groups']
                    diff_indexes = idx_local != idx
                    same_value = self.chart.loc[idx_local, 'wire_terminal_2'] == value_2
                    value_tmp = self.chart.loc[idx_local, 'wire_terminal_1']
                    same_mark = marking.lower() == self.chart.loc[idx_local, MARKING_COL].lower()

                    if diff_indexes and same_value and same_group and not empty(value_tmp) and same_mark:
                        list_1.append(idx)
                        first[idx] = value_tmp
                        # self.chart.loc[idx, 'wire_terminal_1'] = value_tmp
                        break

            if empty(value_2):
                for idx_local, row_local in self.chart.iterrows():
                    same_group = group == self.chart.loc[idx_local, 'groups']
                    diff_indexes = idx_local != idx
                    same_value = value_1 == self.chart.loc[idx_local, 'wire_terminal_1']
                    value_tmp = self.chart.loc[idx_local, 'wire_terminal_2']
                    same_mark = marking.lower() == self.chart.loc[idx_local, MARKING_COL].lower()
                    if diff_indexes and same_value and same_group and not empty(value_tmp) and same_mark:
                        list_2.append(idx)
                        second[idx] = value_tmp
                        # self.chart.loc[idx, 'wire_terminal_2'] = value_tmp
                        break

        for index, row in self.chart.iterrows():
            if index in first:
                self.chart.loc[index, 'wire_terminal_1'] = first[index]
            if index in second:
                self.chart.loc[index, 'wire_terminal_2'] = second[index]

        return (list_1, list_2)

    def __fulfill_NaN(self, terminal):
        return np.nan if terminal is None or terminal == '' or terminal == ' ' else terminal

    def __fulfill_two_cols_empty(self, idxs):
        for idx in idxs:
            marking, group, wire_square, wire_color = self.chart.loc[idx, [
                'marking', 'groups', 'wire_square', 'wire_color'
            ]]
            target_by_group = self.chart.loc[
                (self.chart['marking'] == marking) &
                (self.chart['groups'] == group) &
                (self.chart['wire_terminal_1'] != np.nan) &
                (self.chart['wire_terminal_2'] != np.nan), :]
            if not target_by_group.empty:
                target_by_square = target_by_group.loc[
                    (target_by_group['wire_square'] == wire_square) &
                    (target_by_group['wire_terminal_1'] != np.nan) &
                    (target_by_group['wire_terminal_2'] != np.nan), :]
                if not target_by_square.empty:
                    target_by_color = target_by_square.loc[
                        (target_by_square['wire_color'] == wire_color) &
                        (target_by_square['wire_terminal_1'] != np.nan) &
                        (target_by_square['wire_terminal_2'] != np.nan), :]
                    if not target_by_color.empty:
                        self.chart.loc[idx, ['wire_terminal_1', 'wire_terminal_2']] = target_by_color[
                            ['wire_terminal_1', 'wire_terminal_2']
                        ].iloc[0]
                        print('572')
                    else:
                        self.chart.loc[idx, ['wire_terminal_1', 'wire_terminal_2']] = target_by_square[
                            ['wire_terminal_1', 'wire_terminal_2']
                        ].iloc[0]
                        print('577', target_by_square[
                            ['wire_terminal_1', 'wire_terminal_2']
                        ].iloc[0])
                else:
                    self.chart.loc[idx, ['wire_terminal_1', 'wire_terminal_2']] = target_by_group[
                        ['wire_terminal_1', 'wire_terminal_2']
                    ].iloc[0]
                    print('584', target_by_group[
                        ['wire_terminal_1', 'wire_terminal_2']
                    ].iloc[0])
            else:
                target_by_group = self.chart.loc[
                    (self.chart['marking'] == marking) &
                    (self.chart['groups'] == group) &
                    (self.chart['wire_terminal_1'] != np.nan) | (self.chart['wire_terminal_2'] != np.nan), :]
                if target_by_group:
                    target_by_square = target_by_group.loc[
                        (target_by_group['wire_square'] == wire_square) &
                        ((not empty(target_by_group['wire_terminal_1'])) |
                         (not empty(target_by_group['wire_terminal_2']))), :]
                    if target_by_square:
                        target_by_color = target_by_square[
                            (target_by_square['wire_color'] == wire_color) &
                            ((self.chart['wire_terminal_1'] != np.nan) |
                             (self.chart['wire_terminal_2'] != np.nan)), :]
                        if target_by_color:
                            self.chart.loc[idx, ['wire_terminal_1', 'wire_terminal_2']] = target_by_color[
                                ['wire_terminal_1', 'wire_terminal_2']
                            ].iloc[0]
                            print('606', target_by_color[
                                ['wire_terminal_1', 'wire_terminal_2']
                            ].iloc[0])
                        else:
                            self.chart.loc[idx, ['wire_terminal_1', 'wire_terminal_2']] = target_by_square[
                                ['wire_terminal_1', 'wire_terminal_2']
                            ].iloc[0]
                            print('613', target_by_square[
                                ['wire_terminal_1', 'wire_terminal_2']
                            ].iloc[0])
                    else:
                        self.chart.loc[idx, ['wire_terminal_1', 'wire_terminal_2']] = target_by_group[
                            ['wire_terminal_1', 'wire_terminal_2']
                        ].iloc[0]
                        print('620', target_by_group[
                            ['wire_terminal_1', 'wire_terminal_2']
                        ].iloc[0])
            print('623', self.chart.loc[idx, ['wire_terminal_1', 'wire_terminal_2']])
        return idxs

    def __fulfill_one_col_empty(self, idxs, column_name='wire_terminal_1'):
        for idx in idxs:
            marking, group, wire_square, wire_color, wire_terminal = self.chart.loc[idx, [
                                                                                        'marking',
                                                                                        'groups',
                                                                                        'wire_square',
                                                                                        'wire_color',
                                                                                        column_name,
                                                                                    ]]
            target_by_group = self.chart.loc[
                (self.chart['marking'] == marking) &
                (self.chart['groups'] == group) &
                (self.chart[column_name] != np.nan), :]
            if not target_by_group.empty:
                target_by_square = target_by_group.loc[
                    (target_by_group['wire_square'] == wire_square) &
                    (self.chart[column_name] != np.nan), :]
                if not target_by_square.empty:
                    target_by_color = target_by_square.loc[
                        (target_by_square['wire_color'] == wire_color) &
                        (self.chart[column_name] != np.nan), :]
                    if not target_by_color.empty:
                        self.chart.loc[idx, column_name] = target_by_color[column_name].iloc[0]
                        print('649', target_by_color[[column_name]].iloc[0])
                    else:
                        self.chart.loc[idx, column_name] = target_by_square[column_name].iloc[0]
                        print('652', target_by_square[[column_name]].iloc[0])
                else:
                    self.chart.loc[idx, column_name] = target_by_group[column_name].iloc[0]
                    print('655', target_by_group[column_name].iloc[0])
                print('656', self.chart.loc[idx, [column_name]])

    def __fulfill_terminals_built_in(self):
        self.chart['wire_terminal_1'] = self.chart['wire_terminal_1'].apply(
            lambda x: self.__fulfill_NaN(x)
        )
        self.chart['wire_terminal_2'] = self.chart['wire_terminal_2'].apply(
            lambda x: self.__fulfill_NaN(x)
        )
        empty_terminals_idxs = np.where(pd.isnull(self.chart[['wire_terminal_1', 'wire_terminal_2']]))

        idxs_1 = list()
        idxs_2 = list()
        len_empty_idxs = len(empty_terminals_idxs[0])
        for i in range(len_empty_idxs):
            if empty_terminals_idxs[1][i] == 0:
                idxs_1.append(empty_terminals_idxs[0][i])
            else:
                idxs_2.append(empty_terminals_idxs[0][i])

        two_cols_idxs = self.__fulfill_two_cols_empty(set(idxs_1).intersection(set(idxs_2)))
        self.__fulfill_one_col_empty(set(idxs_1) - two_cols_idxs, column_name='wire_terminal_1')
        self.__fulfill_one_col_empty(set(idxs_2) - two_cols_idxs, column_name='wire_terminal_2')

        for i in range(1, 4):
            str_i = str(i)
            self.chart[(self.chart['marking'] == 'Чёрный') & (self.chart['groups'] == str_i)][['wire_terminal_1', 'wire_terminal_2']] = \
                self.chart[(self.chart['marking'] == 'Чёрный') & (self.chart['groups'] == str_i)][['wire_terminal_1', 'wire_terminal_2']].fillna(method='ffill')
            self.chart[(self.chart['marking'] == 'Чёрный') & (self.chart['groups'] == str_i)][['wire_terminal_1', 'wire_terminal_2']] = \
                self.chart[(self.chart['marking'] == 'Чёрный') & (self.chart['groups'] == str_i)][['wire_terminal_1', 'wire_terminal_2']].fillna(method='bfill')
            self.chart[(self.chart['marking'] == 'Белый') & (self.chart['groups'] == str_i)][['wire_terminal_1', 'wire_terminal_2']] = \
                self.chart[(self.chart['marking'] == 'Белый') & (self.chart['groups'] == str_i)][['wire_terminal_1', 'wire_terminal_2']].fillna(method='ffill')
            self.chart[(self.chart['marking'] == 'Белый') & (self.chart['groups'] == str_i)][['wire_terminal_1', 'wire_terminal_2']] = \
                self.chart[(self.chart['marking'] == 'Белый') & (self.chart['groups'] == str_i)][['wire_terminal_1', 'wire_terminal_2']].fillna(method='bfill')

        return idxs_1, idxs_2

    def __swap_cols_after(self, col1, col2):
        rows_to_swap = []

        for index, row in self.chart.iterrows():
            ax_value = self.chart.loc[index, col2]
            if ax_value is None or ax_value == '' or ax_value == ' ':
                main_value = self.chart.loc[index, col1]
                for idx, row in self.chart.iterrows():
                    if idx > index:
                        ax_value_tmp = self.chart.loc[idx, col2]
                        if ax_value_tmp == main_value:
                            rows_to_swap.append(index)

        self.__swap_sides(rows_to_swap)

    def __reset_index(self):
        self.chart.index = pd.Index(range(self.chart.shape[0]))

    def __delete_cols(self, idx, *cols):
        list_cols = [col for col in cols]
        list_empties = ['' for col in cols]
        self.chart.loc[idx, list_cols] = list_empties

    def __check_for_availability(self, idx, terminal_info, seal_info, terminal_col, seal_col):
        if seal_info:
            if terminal_info:
                if seal_info.seal_available:
                    if terminal_info.terminal_available:
                        if terminal_info.seal_installed:
                            pass
                        else:
                            self.__delete_cols(idx, terminal_col, seal_col)
                    else:
                        self.__delete_cols(idx, terminal_col)
                else:
                    self.__delete_cols(idx, terminal_col, seal_col)
            else:
                if seal_info.seal_available:
                    self.__delete_cols(idx, terminal_col)
                else:
                    self.__delete_cols(idx, terminal_col, seal_col)
        else:
            if terminal_info and terminal_info.terminal_available and not terminal_info.seal_installed:
                self.__delete_cols(idx, seal_col)
            else:
                self.__delete_cols(idx, terminal_col, seal_col)

    def filter_availability_komax_terminal_seal(self, terminals, seals):
        """
        :param terminals_df:
        :return:
        """

        for idx, row in self.chart.iterrows():
            armirovka_1, armirovka_2 = self.chart.loc[idx, [self.ARMIROVKA_1_COL, self.ARMIROVKA_2_COL]]
            terminal_1, terminal_2 = self.chart.loc[idx, [self.TERMINAL_1_COL, self.TERMINAL_2_COL]]
            seal_1, seal_2 = self.chart.loc[idx, [self.SEAL_1_COL, self.SEAL_2_COL]]


            if not is_empty(armirovka_1):
                self.__delete_cols(idx, self.ARMIROVKA_1_COL, self.TERMINAL_1_COL, self.SEAL_1_COL)
            else:
                terminal_1_info, seal_1_info = None, None
                if not is_empty(terminal_1):
                    terminal_1_info = terminals.filter(terminal_name=terminal_1).first()
                if not is_empty(seal_1):
                    seal_1_info = seals.filter(seal_name=seal_1).first()

                self.__check_for_availability(
                    idx,
                    terminal_1_info,
                    seal_1_info,
                    self.TERMINAL_1_COL,
                    self.SEAL_1_COL,
                )

            if not is_empty(armirovka_2):
                self.__delete_cols(idx, self.ARMIROVKA_2_COL, self.TERMINAL_2_COL, self.SEAL_2_COL)
            else:
                terminal_2_info, seal_2_info = None, None
                if not is_empty(terminal_2):
                    terminal_2_info = terminals.filter(terminal_name=terminal_2).first()
                if not is_empty(seal_2):
                    seal_2_info = seals.filter(seal_name=seal_2).first()

                self.__check_for_availability(
                    idx,
                    terminal_2_info,
                    seal_2_info,
                    self.TERMINAL_2_COL,
                    self.SEAL_2_COL,
                )

    def sort(self, method='simple', first_sort=False, test=False):

        self.delete_word_contain('*')

        self.__correct_marking()

        self.chart.sort_values(
            by=['marking'],
            ascending=True,
            inplace=True
        )

        self.__reset_index()

        if test:
            self.chart.to_excel('test0.xlsx')


        self.__swap_cols('wire_terminal_1', 'wire_terminal_2')

        self.__reset_index()

        self.__divide_in_groups()

        pairing_required = str(int('wire_terminal_3' in self.chart.columns))

        if test:
            self.chart.to_excel('test.xlsx')

        #TODO: increase speed of func

        # idxs_1, idxs_2 = self.__fulfill_terminals()
        idxs_1, idxs_2 = self.__fulfill_terminals_built_in()

        if test:
            self.chart.to_excel('test1.xlsx')

        if first_sort:
            self.chart.sort_values(
                by=['wire_type', 'marking', 'groups', 'wire_terminal_1', 'wire_terminal_2', 'wire_square', 'wire_color'],
                ascending=True,
                inplace=True
            )
        else:
            self.chart.sort_values(
                by=['wire_type', 'marking', 'groups', 'wire_terminal_2', 'wire_terminal_1', 'wire_square', 'wire_color'],
                ascending=True,
                inplace=True
            )


        if test:
            self.chart.to_excel('test2.xlsx')

        for idx in idxs_1:
            self.chart.loc[idx, 'wire_terminal_1'] = ''
        for idx in idxs_2:
            self.chart.loc[idx, 'wire_terminal_2'] = ''

        self.__reset_index()

        if method == 'double':
            # self.__swap_cols_after('wire_terminal_1', 'wire_terminal_2')

            idxs_1, idxs_2 = self.__fulfill_terminals()

            if first_sort:
                self.chart.sort_values(
                    by=['wire_type', 'marking', 'groups', 'wire_terminal_1', 'wire_terminal_2', 'wire_square', 'wire_color'],
                    ascending=True,
                    inplace=True
                )
            else:
                self.chart.sort_values(
                    by=['wire_type', 'marking', 'groups', 'wire_terminal_2', 'wire_terminal_1', 'wire_square', 'wire_color'],
                    ascending=True,
                    inplace=True
                )

            self.__reset_index()

            if test:
                self.chart.to_excel('test.xlsx')

            for idx in idxs_1:
                self.chart.loc[idx, 'wire_terminal_1'] = ''
            for idx in idxs_2:
                self.chart.loc[idx, 'wire_terminal_2'] = ''

            # self.__swap_cols_gaps('wire_terminal_2', 'wire_terminal_1')

        """
        self.chart.drop(
            columns=['groups'],
            inplace=True
        )
        """

        # self.__reset_index()

    def __time_changeover(self, idx, time, volume=1, last_first=None, last_second=None, pairing=False, full=False,
                          last_index=None, swapped=False):
        square_change, color_change, conn1_change, conn2_change = False, False, False, False
        armirovka1_change, armirovka2_change = False, False
        time_position = 0

        if full:
            return time['learn'] * 3 + time['aplicator'] * 2 + time['direction'] + time['contact'] * 2 + time['wire'] +\
                   time['compact'] + time['pack'] + time['ticket'] + time['task'] + time['cut'] * volume

        if not pairing:
            if last_index is not None and swapped:
                self.__swap_sides([idx], copy=True)

            cols = list(self.chart)
            for col in cols:
                value = self.chart.loc[idx, col]
                if idx > 0:
                    if last_index is None:
                        back_value = self.chart.loc[idx - 1, col]
                    else:
                        back_value = self.new_chart.loc[last_index, col]

                    if value != back_value:
                        if col == 'wire_color':
                            color_change = True
                        elif col == 'wire_square':
                            square_change = True
                        elif col == 'wire_terminal_1':
                            if value is not None and value != '' and value != ' ':
                                if back_value is None or back_value == '' or back_value == ' ':
                                    if last_first is not None and last_first == value:
                                        conn1_change = False
                                    else:
                                        conn1_change = True
                                else:
                                    conn1_change = True
                            else:
                                conn1_change = False
                        elif col == 'wire_terminal_2':
                            if value is not None and value != '' and value != ' ':
                                if back_value is None or back_value == '' or back_value == ' ':
                                    if last_second is not None and last_second == value:
                                        conn2_change = False
                                    else:
                                        conn2_change = True
                                else:
                                    conn2_change = True
                            else:
                                conn2_change = False
                        elif col == 'wire_seal_1':
                            if value is not None and value != '' and value != ' ':
                                if back_value is None or back_value == '' or back_value == ' ':
                                    if conn1_change:
                                        armirovka1_change = True
                                    else:
                                        armirovka1_change = False
                                else:
                                    armirovka1_change = True
                            else:
                                armirovka1_change = False
                        elif col == 'wire_seal_2' and value is not None and value != '':
                            if value is not None and value != '' and value != ' ':
                                if back_value is None or back_value == '' or back_value == ' ':
                                    if conn2_change:
                                        armirovka2_change = True
                                    else:
                                        armirovka2_change = False
                                else:
                                    armirovka2_change = True
                            else:
                                armirovka2_change = False
                else:
                    color_change, square_change, conn1_change, conn2_change = True, True, True, True
                    armirovka2_change, armirovka1_change = True, True

            if last_index is not None and swapped:
                self.__swap_sides([idx], copy=True)

            if square_change:
                if conn1_change and conn2_change:
                    time_position += (time['learn'] * 3 + time['aplicator'] * 2 + time['direction'] + time['contact'])
                elif not conn1_change and not conn2_change:
                    time_position += (time['learn'] + time['direction'])
                else:
                    time_position += (time['learn'] * 2 + time['aplicator'] + time['direction'] + time['contact'])
            else:
                if conn1_change and conn2_change:
                    time_position += (time['learn'] * 2 + time['aplicator'] * 2 + time['contact'] * 2)
                elif not conn1_change and not conn2_change:
                    pass
                else:
                    time_position += (time['learn'] + time['aplicator'] + time['contact'])

            if square_change or color_change:
                time_position += time['wire']

            if square_change:
                time_position += time['compact']

            if armirovka1_change:
                time_position += time['compact']
            if armirovka2_change:
                time_position += time['compact']


            time_position += (time['pack'] + time['ticket'] + time['task'])

            time_position += volume * time['cut']
        else:
            time_position = time['aplicator']

        return time_position

    def __is_marking_black(self, marking):
        marking = marking.lower()
        return 'черный' in marking or 'чёрный' in marking

    def __is_marking_white(self, marking):
        return 'белый' in marking.lower()

    def __get_last_value_of_column(self, current_index, column_name):
        """
        Finds last(nearest) value from column(especially from terminal columns)

        :param current_index: idxd of current row
        :param column_name: name of terminal column
        :return: value of last(nearest) cell
        """

        back_col_value = self.chart.loc[current_index - 1, column_name]
        col_value = self.chart.loc[current_index, column_name]

        last_value = None, None

        if back_col_value is None or back_col_value == ' ' or back_col_value == '':
            if col_value is not None and col_value != ' ' and col_value != '':
                index = current_index - 1
                while index > 0:
                    if back_col_value is not None and back_col_value != ' ' and back_col_value != '':
                        break
                    else:
                        index -= 1
                    back_col_value = self.chart.loc[index, column_name]
                last_value = self.chart.loc[index, column_name]

        return last_value

    def __time_quantity_allocation(self, quantity, time):
        """
        count time on each row
        change self.chart

        :param quantity: dict, harness_number(str): quantity(int)
        :param time: labourisness dict, action(str): sec(int)
        :param hours: shift time, int
        :return:
        """


        first_black_passed = False
        first_white_passed = False

        if not self.AMOUNT_COL in self.chart:
            self.chart[self.AMOUNT_COL] = 0

        for idx, row in self.chart.iterrows():
            # harness number inserting
            harness_number = self.chart.loc[idx, self.HARNESS_NUMBER_COL]
            amount = self.chart.loc[idx, self.AMOUNT_COL]

            """
            print(curr_amount)
            if curr_amount > 1:
                pass
            else:
                if type(quantity) is dict:
                    amount = quantity[harness_number]
                else:
                    amount = quantity
            """

            # time inserting
            marking = self.chart.loc[idx, self.MARKING_COL]
            if self.__is_marking_black(marking) and not first_black_passed:
                first_black_passed = True
                time_changeover_row = self.__time_changeover(idx, time, amount, full=True)
            elif self.__is_marking_white(marking) and not first_white_passed:
                first_white_passed = True
                time_changeover_row = self.__time_changeover(idx, time, amount, full=True)
            else:
                last_first = self.__get_last_value_of_column(idx, self.TERMINAL_1_COL)
                last_second = self.__get_last_value_of_column(idx, self.TERMINAL_2_COL)

                time_changeover_row = self.__time_changeover(idx, time, amount, last_first, last_second)

            self.chart.loc[idx, self.TIME_COL] = time_changeover_row

    def __get_allocated_time(self, quantity, time):
        self.__time_quantity_allocation(quantity, time)
        return sum(self.chart.loc[:, 'time'])

    def make_best_sort(self, quantity, time):
        self.sort(method='simple', first_sort=True)
        first_time = self.__get_allocated_time(quantity, time)
        self.sort(method='simple', first_sort=False)
        second_time = self.__get_allocated_time(quantity, time)

        if second_time > first_time:
            pass
        else:
            self.sort(method='simple', first_sort=True)

        return

    def __consistently_allocation(self, komaxes, kappas, quantity, time, hours, group_of_square, allocation,
                                  returned='error'):
        """
        allocation pass by, NOT parallel

        if not enough komaxes without pairing ability, gets komaxes with pairing ability
        if not enough komaxes with pairing ability, gets komaxes without pairing ability
        
        :param: returned, error if alloc > hours(shift) or allocation if alloc without error
        :param komaxes: dict, komax(int): [status(int:1, 0), marking(int: 1(white), 2(black)), pairing(int: 1, 0)]
        :param quantity: dict, harness_number(str): amount(int)
        :param time: labourisness dict, acting(str): sec(int/float)
        :param hours: shift in hours, int
        :return: allocation dict, komax_number(int): seconds(float)
        """
        
        if returned == 'error':
            returned = 'None'
        elif returned == 'allocation':
            returned = 'komax'
            
        self.__correct_marking()
        self.__time_quantity_allocation(quantity, time)

        # create dict of allocation
        alloc = allocation
        shift = hours*60*60       # in seconds

        # get komax config info
        black_pairing_komax, black_komax, white_pairing_komax, white_komax = self.__get_black_white_komax_key(komaxes)
        amount_black_pairing_komax, amount_white_pairing_komax, \
        amount_black_komax, amount_white_komax = self.__get_amount_black_white_komax(komaxes)

        black_pairing_komax_more = amount_black_pairing_komax >= amount_white_pairing_komax
        black_komax_more = amount_black_komax >= amount_white_komax
        
        # error indicator
        error = False

        # without pairing komaxes
        for idx, row in self.chart.iterrows():
            wire_type = self.chart.loc[idx, self.WIRE_TYPE_COL]
            if type(wire_type) is str and 'Кабель' in wire_type:
                if len(kappas) and kappas[0] is not None:
                    self.chart.loc[idx, self.KAPPA_COL] = kappas[0].number
                else:
                    self.chart.loc[idx, self.KAPPA_COL] = None
                continue

            marking = self.chart.loc[idx, self.MARKING_COL]
            time_changeover_position = self.chart.loc[idx, self.TIME_COL]
            square_group = group_by(self.chart.loc[idx, self.SQUARE_COL])
            if self.__is_marking_black(marking) and square_group == group_of_square and black_komax != 0:
                alloc[black_komax][0] += time_changeover_position
                self.chart.loc[idx, self.KOMAX_COL] = black_komax
            elif self.__is_marking_white(marking) and square_group == group_of_square and white_komax != 0:
                alloc[white_komax][0] += time_changeover_position
                self.chart.loc[idx, self.KOMAX_COL] = white_komax
            else:
                if square_group != group_of_square and square_group is not None:
                    pass
                else:
                    error = True

            if black_komax is not None and black_komax != 0 and alloc[black_komax][0] >= shift:
                black_komax = self.__get_next_komax(alloc, shift, komaxes, black_komax, 1, 1, 0, group_of_square, 
                                                    returned)
            elif white_komax is not None and white_komax != 0 and alloc[white_komax][0] >= shift:
                white_komax = self.__get_next_komax(alloc, shift, komaxes, white_komax, 1, 2, 0, group_of_square, 
                                                    returned)
            if black_komax is None or white_komax is None:
                return -1

        return (alloc, error)

    def __get_black_white_komax_key(self, komaxes):
        black_pairing_komax = 0
        white_pairing_komax = 0
        black_komax = 0
        white_komax = 0

        for key, item in komaxes.items():
            if komaxes[key][0] == 1:
                if komaxes[key][1] == 2:
                    if komaxes[key][2] == 1 and white_pairing_komax == 0:
                        white_pairing_komax = key
                    elif komaxes[key][2] == 0 and white_komax == 0:
                        white_komax = key
                elif komaxes[key][1] == 1:
                    if komaxes[key][2] == 1 and black_pairing_komax == 0:
                        black_pairing_komax = key
                    elif komaxes[key][2] == 0 and black_komax == 0:
                        black_komax = key

        if black_komax == 0:
            black_komax = black_pairing_komax
        if white_komax == 0:
            white_komax = white_pairing_komax

        return (black_pairing_komax, black_komax, white_pairing_komax, white_komax)

    def __get_amount_black_white_komax(self, komaxes):
        amount_black_pairing_komax = 0
        amount_white_pairing_komax = 0

        amount_black_komax = 0
        amount_white_komax = 0

        for key, item in komaxes.items():
            if item[0] == 1 and item[1] == 1 and item[2] == 1:
                amount_white_pairing_komax += 1
            elif item[0] == 1 and item[1] == 2 and item[2] == 1:
                amount_black_pairing_komax += 1
            elif item[0] == 1 and item[1] == 1 and item[2] == 0:
                amount_white_komax += 1
            elif item[0] == 1 and item[1] == 2 and item[2] == 0:
                amount_black_komax += 1

        return (amount_black_pairing_komax, amount_black_komax, amount_white_pairing_komax, amount_white_komax)

    def __get_next_komax(self, alloc, shift, komaxes, current_komax, status=1, marking=1, pairing=0, square_group=1, 
                         returned='None'):
        """
        Get komax with suitable params with index more than current
        
        :param: output, None or komax, if current komax need to be returned
        :param current_komax: index of current komax
        :param komaxes: komaxes dict
        :return: index of next working komax of type current_komax or None
        """

        for komax_idx, params in komaxes.items():
            if komax_idx > current_komax and status == params[0] and marking == params[1] and pairing == params[2] and \
                    square_group in params[3]:
                return komax_idx

        for komax_idx, params in komaxes.items():
            if alloc[komax_idx][0] <= shift and status == params[0] and marking == params[1] and \
                    int(not bool(pairing)) == params[2] and square_group in params[3]:
                return komax_idx
        
        if returned == 'None':
            return None
        elif returned == 'komax':
            return current_komax

    def __get_all_komaxes_keys(self, komaxes, status, marking, pairing):
        output_komaxes = list()

        for komax_idx, params in komaxes.items():
            if type(pairing) is list:
                if status == params[0] and marking == params[1] and (pairing[0] == params[2] or pairing[1] == params[2]):
                    output_komaxes.append(komax_idx)
            elif type(pairing) is int:
                if status == params[0] and marking == params[1] and pairing == params[2]:
                    output_komaxes.append(komax_idx)

        return output_komaxes

    def __get_sum_black_white_marking_time(self, group_of_square):
        sum_black_marking, sum_white_marking = 0, 0

        for idx, row in self.chart.iterrows():
            marking = self.chart.loc[idx, self.MARKING_COL]
            time = self.chart.loc[idx, self.TIME_COL]
            square = self.chart.loc[idx, self.SQUARE_COL]
            group = group_by(square)

            if self.__is_marking_black(marking) and group_of_square == group:
                sum_black_marking += time
            elif self.__is_marking_white(marking) and group_of_square == group:
                sum_white_marking += time
            else:
                pass

        return (sum_black_marking, sum_white_marking)

    def average(self, items: int, divider: int):
        if divider == 0 and items == 0:
            return 0
        elif divider == 0 and items != 0:
            return None
        else:
            return items / divider

    def __create_komax_col(self):
        self.chart[self.KOMAX_COL] = None

    def __create_kappa_col(self):
        self.chart[self.KAPPA_COL] = None

    def __parallel_allocation(self, komaxes, kappas, quantity, time, hours, group_of_square, allocation,
                              returned='error'):
        """
        parallel allocation

        Uses all komaxes without priority
        
        :param: returned, error if alloc > hours(shift) or allocation if alloc without error
        :param komaxes: dict, komax(int): [status(int:1, 0), marking(int: 1(white), 2(black)), pairing(int: 1, 0)]
        :param quantity: dict, harness_number(str): amount(int)
        :param time: labourisness dict, acting(str): sec(int/float)
        :param hours: shift in hours, int
        :return: allocation dict, komax_number(int): seconds(float)
        """

        self.__correct_marking()
        self.__time_quantity_allocation(quantity, time)

        # create dict of allocation
        alloc = allocation
        
        # errors indicator
        error = False
        
        black_komaxes = self.__get_all_komaxes_keys(komaxes, 1, 1, [0, 1])
        amount_black_komaxes = len(black_komaxes)
        white_komaxes = self.__get_all_komaxes_keys(komaxes, 1, 2, [0, 1])
        amount_white_komaxes = len(white_komaxes)
        black_komax_curr_idx = 0
        white_komax_curr_idx = 0

        sum_black_marking_time, sum_white_marking_time = self.__get_sum_black_white_marking_time(group_of_square)

        average_black_time = self.average(sum_black_marking_time, len(black_komaxes))
        average_white_time = self.average(sum_white_marking_time, len(white_komaxes))

        if average_black_time is None or average_white_time is None:
            error = True
            return (alloc, error)

        # without pairing komaxes
        for idx, row in self.chart.iterrows():

            wire_type = self.chart.loc[idx, self.WIRE_TYPE_COL]
            if type(wire_type) is str and 'Кабель' in wire_type:
                if len(kappas) and kappas[0] is not None:
                    self.chart.loc[idx, self.KAPPA_COL] = kappas[0].number
                else:
                    self.chart.loc[idx, self.KAPPA_COL] = None
                continue

            marking = self.chart.loc[idx, self.MARKING_COL]
            time_changeover_position = self.chart.loc[idx, self.TIME_COL]
            group_square = group_by(self.chart.loc[idx, self.SQUARE_COL])
            if self.__is_marking_black(marking) and amount_black_komaxes > 0 and group_square == group_of_square:
                alloc[black_komaxes[black_komax_curr_idx]][0] += time_changeover_position
                self.chart.loc[idx, self.KOMAX_COL] = black_komaxes[black_komax_curr_idx]
            elif self.__is_marking_white(marking) and amount_white_komaxes > 0 and group_square == group_of_square:
                alloc[white_komaxes[white_komax_curr_idx]][0] += time_changeover_position
                self.chart.loc[idx, self.KOMAX_COL] = white_komaxes[white_komax_curr_idx]
            else:
                if group_square != group_of_square and group_square is not None:
                    pass
                else:
                    error = True
            if black_komax_curr_idx < amount_black_komaxes and len(black_komaxes) and \
                    alloc[black_komaxes[black_komax_curr_idx]][0] > average_black_time:
                black_komax_curr_idx += 1
            elif white_komax_curr_idx < amount_white_komaxes and len(white_komaxes) and \
                    alloc[white_komaxes[white_komax_curr_idx]][0] > average_white_time:
                white_komax_curr_idx += 1
            
            if white_komax_curr_idx == amount_white_komaxes:
                if returned == 'error':
                    return -1
                elif returned == 'allocation':
                    white_komax_curr_idx -= 1
            if black_komax_curr_idx == amount_black_komaxes:
                if returned == 'error':
                    return -1
                elif returned == 'allocation':
                    black_komax_curr_idx -= 1

        return (alloc, error)

    def allocate(self, komaxes, kappas, quantity, time, hours=None, type_of_allocation='parallel'):
        komax_loading = KomaxLoading(komaxes)
        self.__create_komax_col()
        self.__create_kappa_col()
        groups = (1, 2, 3)
        for group in groups:
            komaxes_group, allocation_group = komax_loading.get_komaxes_allocation_by_komax_group_square(group)
            if type_of_allocation == 'parallel':
                alloc_for_group, error = self.__parallel_allocation(komaxes_group,
                                                                    kappas,
                                                                    quantity, 
                                                                    time, 
                                                                    hours, 
                                                                    group, 
                                                                    allocation_group,
                                                                    returned='allocation')
                # print(alloc_for_group, error)
            elif type_of_allocation == 'consistently':
                alloc_for_group, error = self.__consistently_allocation(komaxes_group,
                                                                        kappas,
                                                                        quantity, 
                                                                        time, 
                                                                        hours, 
                                                                        group, 
                                                                        allocation_group,
                                                                        returned='allocation')
                # print(alloc_for_group, error)
            else:
                return -1
            
            komax_loading.set_allocation(alloc_for_group)

            if error:
                return -1

        return komax_loading.get_allocation_dict()
        
    def smart_sort(self, time, amount):
        HARNESS_NUMBER_COL = "harness"
        base_rows = self.chart_copy.shape[0]
        sum_time = 0
        min_time = 0

        self.new_chart['Amount'] = 0

        for base_row in range(base_rows):
            main_row = self.new_chart.shape[0] - 1

            if base_row == 0:
                idx = random.randint(0, base_rows - 1)
                self.new_chart = self.new_chart.append(self.chart_copy.loc[idx, :])
                self.new_chart.index = pd.Index(range(self.new_chart.shape[0]))
                volume = amount[self.chart_copy.loc[idx, HARNESS_NUMBER_COL]]
                self.new_chart.loc[base_row, 'Amount'] = volume
                self.chart_copy.drop(index=idx, inplace=True)
                continue

            if base_row != main_row and main_row >= 0:
                min_time = 10000
                min_index = 0
                to_swap = False
                for index, row in self.chart_copy.iterrows():
                    volume = amount[self.chart_copy.loc[index, HARNESS_NUMBER_COL]]
                    time1 = self.__time_changeover(index, time, volume=volume, last_index=main_row)
                    time2 = self.__time_changeover(index, time, volume=volume, last_index=main_row, swapped=True)



                    if time1 > time2:
                        if min_time > time2:
                            min_time = time2
                            min_index = index
                            to_swap = True
                    else:
                        if min_time > time1:
                            min_time = time1
                            to_swap = False
                            min_index = index

                if to_swap:
                    self.__swap_sides([min_index], copy=True)

                self.new_chart = self.new_chart.append(self.chart_copy.loc[min_index, :])
                self.new_chart.index = pd.Index(range(self.new_chart.shape[0]))
                volume = amount[self.chart_copy.loc[min_index, HARNESS_NUMBER_COL]]
                self.new_chart.loc[base_row, 'Amount'] = volume
                self.chart_copy.drop(index=min_index, inplace=True)

            sum_time += min_time

        return sum_time

    #TODO : big time needed, incorrect algo, too many changeovers
    def task_allocate(self, komaxes, quantity, time, hours):

        """
        required quicksort with params: pairing-marking-square-terminal-terminal

        :param komaxes: {komax_number : komax_status(0, 1), komax_marking(1 - w, 2 - b), komax_pairing (0, 1)
        :param quantity: {harness_number : amount}
        :param time: {ops name : time(sec)}
        :param hours: hours
        :return: {komax_number : loaded_time(sec)}
        """

        HARNESS_NUMBER_COL = "harness"
        KOMAX_COL = 'komax'
        MARKING_COL = "marking"
        HOURS_TO_SEC = 3600

        self.chart[KOMAX_COL] = ''

        alloc = {i : 0 for i in range(1, len(komaxes) + 1)}

        curr_marking = 0
        curr_pairing = 0
        curr_group_terminal = ''
        idx_start = 0
        time_changeover = 0

        first_sort = self.chart.nunique()["wire_terminal_1"] >= self.chart.nunique()["wire_terminal_2"]

        list_groups = {}

        for idx, row in self.chart.iterrows():
            charachteristic_value = row['col_for_sort'].split(' ')

            new_marking = charachteristic_value[1] != curr_marking
            new_pairing = charachteristic_value[0] != curr_pairing
            new_group_terminal = charachteristic_value[3] != curr_group_terminal
            last_idx = idx == len(self.chart)

            curr_marking = 1 if 'белый' in charachteristic_value[1].lower() else 2
            curr_pairing = int(charachteristic_value[0])
            curr_group_terminal = charachteristic_value[3]

            volume = quantity[self.chart.loc[idx, HARNESS_NUMBER_COL]]

            if new_marking or new_pairing or new_group_terminal:
                if not last_idx:
                    list_groups[idx_start] = [curr_pairing, curr_marking, time_changeover, idx]
                else:
                    time_changeover += self.__time_changeover(idx, time, volume, curr_pairing)
                    list_groups[idx_start] = [curr_pairing, curr_marking, time_changeover, idx + 1]

                time_changeover = 0

                idx_start = idx
                time_changeover += self.__time_changeover(idx, time, volume, curr_pairing, full=True)
            else:
                time_changeover += self.__time_changeover(idx, time, volume, curr_pairing)

        for key, item in list_groups.items():
            for komax_num, komax_set in komaxes.items():
                has_free_time = hours * HOURS_TO_SEC - alloc[komax_num] > item[2]
                pairing = komax_set[2] == item[0]
                marking = komax_set[1] == item[1]
                working = komax_set[0] != 0

                if has_free_time and pairing and marking and working:
                    idx_start = key
                    idx_end = item[3]
                    for idx in range(idx_start, idx_end):
                        self.chart.loc[idx, KOMAX_COL] = komax_num
                    alloc[komax_num] += float(item[2])
                elif not has_free_time and pairing and marking and working:
                    free_time_all = 0
                    komax_set = {}
                    for komax_num_local, komax_set_local in komaxes.items():
                        has_free_time = hours * HOURS_TO_SEC - alloc[komax_num_local] > item[2]
                        pairing = komax_set_local[2] == item[0]
                        marking = komax_set_local[1] == item[1]
                        working = komax_set_local[0] != 0
                        if has_free_time and pairing and marking and working:
                            free_time_local = hours * HOURS_TO_SEC - alloc[komax_num_local]
                            if komax_num_local in komax_set.keys():
                                komax_set[komax_num_local] += free_time_local
                            else:
                                komax_set[komax_num_local] = free_time_local
                            free_time_all += hours * HOURS_TO_SEC - alloc[komax_num_local]

                    if free_time_all < item[2]:
                        # means that for this group free time of all komaxes less than required to cut this group
                        return -1
                    else:
                        idx_start_local = key
                        idx_end_local = item[3]
                        for komax_num_local, komax_free_time in komax_set.items():
                            # better get last row and count time of changeover better, but
                            # will count as full

                            for idx_local in range(idx_start_local, idx_end_local):
                                volume = quantity[self.chart.loc[idx_local, HARNESS_NUMBER_COL]]
                                time_changeover_local = 0
                                if idx_local == idx_start_local:
                                    time_changeover_local = self.__time_changeover(idx_local, time, volume, pairing, True)
                                else:
                                    time_changeover_local = self.__time_changeover(idx_local, time, volume, pairing)

                                komax_free_time -= time_changeover_local
                                if komax_free_time > 0:
                                    self.chart.loc[idx_local, KOMAX_COL] = komax_num_local
                                else:
                                    komax_free_time += time_changeover_local
                                    idx_start_local = idx_local
                                    break

                        if self.chart.loc[idx_end_local - 1] == '':
                            # means that end of komaxes
                            return -2

        return alloc

    #TODO : incorrect changeover counting
    def task_allocation(self, komaxes, quantity, time, hours):

        self.__correct_marking()

        HARNESS_NUMBER_COL = "harness"
        AMOUNT_COL = "amount"
        KOMAX_COL = 'komax'
        MARKING_COL = "marking"
        TERMINAL_1_COL = 'wire_terminal_1'
        TERMINAL_2_COL = 'wire_terminal_2'

        """
        try:
            length2 = get_index(df, "Длина, мм (± 3мм) 2")[0]
        except TypeError:
            length2 = -1
        """

        # create dict of allocation
        alloc = {i: [0] for i in komaxes}
        worker_time = hours*60*60       # in seconds

        # get info of komaxes

        black_pairing_komax = 0
        white_pairing_komax = 0
        black_komax = 0
        white_komax = 0

        for key, item in komaxes.items():
            if komaxes[key][0] == 1:
                if komaxes[key][1] == 2:
                    if komaxes[key][2] == 1 and white_pairing_komax == 0:
                        white_pairing_komax = key
                    elif komaxes[key][2] == 0 and white_komax == 0:
                        white_komax = key
                elif komaxes[key][1] == 1:
                    if komaxes[key][2] == 1 and black_pairing_komax == 0:
                        black_pairing_komax = key
                    elif komaxes[key][2] == 0 and black_komax == 0:
                        black_komax = key

        if black_komax == 0:
            black_komax = black_pairing_komax
        if white_komax == 0:
            white_komax = white_pairing_komax

        amount_black_pairing_komax = 0
        amount_white_pairing_komax = 0

        amount_black_komax = 0
        amount_white_komax = 0

        for key, item in komaxes.items():
            if item[0] == 1 and item[1] == 1 and item[2] == 1:
                amount_white_pairing_komax += 1
            elif item[0] == 1 and item[1] == 2 and item[2] == 1:
                amount_black_pairing_komax += 1
            elif item[0] == 1 and item[1] == 1 and item[2] == 0:
                amount_white_komax += 1
            elif item[0] == 1 and item[1] == 2 and item[2] == 0:
                amount_black_komax += 1

        black_pairing_komax_more = amount_black_pairing_komax >= amount_white_pairing_komax
        black_komax_more = amount_black_komax >= amount_white_komax

        # adding columns with harnesses amount and komax number alloc
        if quantity is not None:
            self.chart[AMOUNT_COL] = ''
        self.chart[KOMAX_COL] = ''

        #count time for paired positions
        """
        if length2 == -1:
            pass
        else:
            time_spend_black_pairing = 0
            time_spend_white_pairing = 0

            no_more_black_pairing = False
            no_more_white_pairing = False

            for row in range(1, rows):
                value = df.iloc[row, length2]
                quantity_harness = 0
                time_position = 0
                if value is None:
                    pass
                else:
                    harness = df.ix[row, harness_number_col]
                    # quantity_harness = 1
                    try:
                    # quantity might be None
                        quantity_harness = quantity[harness]
                    except KeyError:
                        text_warn_rus_1 = "Проверь номера жгутов в файле кол-ва жгутов."
                        text_warn_rus_2 = "Не найден жгут номер {}".format(harness)
                        write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1 + text_warn_rus_2)
                        raise KeyError("Check if harness number is correct in \"Количество.xlsx\" file! : {}".format(harness))

                    time_position = time_changeover(df, row, column_numbers, time, volume=quantity_harness, pairing=True)

                    df.iloc[row, cols - 2] = quantity_harness

                    marking = df.iloc[row, mark_col]
                    if type(marking) is str and 'б' in marking.lower():
                        if not no_more_white_pairing:
                            time_spend_white_pairing += time_position
                            if white_pairing_komax == 0:
                                text_warn_rus_1 = "Нет комаксов с белой маркировкой и возможностью спарки!"
                                write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                                raise Warning("No komaxes with white marking and pairing ability")
                            df.iloc[row, cols - 1] = white_pairing_komax
                        else:
                            #TODO: replace "спарки" with "спаривание" everywhere please!
                            text_warn_rus_1 = "Слишком много позиций для спарки с белой маркировкой!"
                            write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                            # raise Warning("too many positions for pairing with white marking!")
                    elif type(marking) is str and 'ч' in marking.lower():
                        if not no_more_black_pairing:
                            time_spend_black_pairing += time_position
                            if black_pairing_komax == 0:
                                text_warn_rus_1 = "Нет комаксов с возможностью спарки и чёрной маркировкой!"
                                write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                                raise Warning("No komaxes with black marking and pairing ability")
                            df.iloc[row, cols - 1] = black_pairing_komax
                        else:
                            text_warn_rus_1 = "Слишком много позиций для спарки с чёрной маркировкой!"
                            write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                            # raise Warning("too many positions for pairing with black marking!")
                    elif marking is None:
                        # count whether more white or black paired komaxes
                        if black_pairing_komax_more:
                            if not no_more_black_pairing and black_pairing_komax != 0:
                                time_spend_black_pairing += time_position
                                df.iloc[row, cols - 1] = black_pairing_komax
                            elif not no_more_white_pairing and white_pairing_komax != 0:
                                time_spend_white_pairing += time_position
                                df.iloc[row, cols - 1] = white_pairing_komax
                            else:
                                text_warn_rus_1 = "Слишком много позиций для спарки без маркировки!"
                                text_warn_rus_2 = "Невозможно их распределить по komax"
                                write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1 + text_warn_rus_2)
                                # raise Warning("too many positions for pairing with no marking!")
                        else:
                            if not no_more_white_pairing and white_pairing_komax != 0:
                                time_spend_white_pairing += time_position
                                df.iloc[row, cols - 1] = white_pairing_komax
                            elif not no_more_black_pairing and black_pairing_komax != 0:
                                time_spend_black_pairing += time_position
                                df.iloc[row, cols - 1] = black_pairing_komax

                            else:
                                text_warn_rus_1 = "Слишком много позиций для спарки без маркировки!"
                                text_warn_rus_2 = "Невозможно их распределить по komax"
                                write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1 + text_warn_rus_2)
                                # raise Warning("too many positions for pairing with no marking!")

                    last_row = row == rows - 1
                    if time_spend_black_pairing > worker_time or last_row:
                        if black_pairing_komax == 0:
                            text_warn_rus_1 = "Нет komax с чёрной маркировкой и возможностью спарки!"
                            write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                            raise Warning("No komaxes with black marking and pairing ability at all!")
                        alloc[black_pairing_komax][0] += time_spend_black_pairing
                        if not last_row:
                            for key, item in komaxes.items():
                                komax_work = komaxes[key][0] == 1
                                komax_black = komaxes[key][1] == 2
                                komax_pair = komaxes[key][2] == 1
                                current_komax = key == black_pairing_komax
                                less_worker_time = alloc[key][0] < worker_time
                                if komax_work and komax_black and komax_pair and not current_komax and less_worker_time:
                                    black_pairing_komax = key
                                    time_spend_black_pairing = 0
                                    break
                                else:
                                    no_more_black_pairing = True

                    if time_spend_white_pairing > worker_time or last_row:
                        if white_pairing_komax == 0:
                            text_warn_rus_1 = "Нет комаксов с возможностью спаривания и белой маркировкой!"
                            write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                            raise Warning("No komaxes with white marking and pairing ability at all!")
                        alloc[white_pairing_komax][0] += time_spend_white_pairing
                        if not last_row:
                            for key, item in komaxes.items():
                                komax_work = komaxes[key][0] == 1
                                komax_white = komaxes[key][1] == 1
                                komax_pair = komaxes[key][2] == 1
                                current_komax = key == white_pairing_komax
                                less_worker_time = alloc[key][0] < worker_time
                                if komax_work and komax_white and komax_pair and not current_komax and less_worker_time:
                                    white_pairing_komax = key
                                    time_spend_black_pairing = 0
                                    break
                                else:
                                    no_more_white_pairing = True
        """

        # count time for non-paired positions

        time_spend_black = 0
        time_spend_white = 0

        no_more_black = False
        no_more_white = False

        first_black = False
        first_white = False

        for idx, row in self.chart.iterrows():
            # value = df.iloc[row, length2]
            value = None
            if value is None:
                harness = self.chart.loc[idx, HARNESS_NUMBER_COL]

                if quantity is not None:
                    quantity_harness = quantity[harness]
                else:
                    quantity_harness = self.chart.loc[idx, AMOUNT_COL]
                """
                try:
                    quantity_harness = quantity[harness]
                except KeyError:
                    raise KeyError("Check if harness number is correct in \"quantity.xlsx\" file! : {}".format(harness))
                """

                if ('черный' in self.chart.loc[idx, MARKING_COL].lower() or 'чёрный' in self.chart.loc[idx, MARKING_COL].lower()) and not first_black:
                    first_black = True
                    time_position = self.__time_changeover(idx, time, volume=quantity_harness, full=True)
                elif 'белый' in self.chart.loc[idx, MARKING_COL].lower() and not first_white:
                    first_white = True
                    time_position = self.__time_changeover(idx, time, volume=quantity_harness, full=True)
                else:
                    #TODO: Try to make func for this part
                    back_first_col_value = self.chart.loc[idx - 1, TERMINAL_1_COL]
                    first_col_value = self.chart.loc[idx, TERMINAL_1_COL]

                    back_second_col_value = self.chart.loc[idx - 1, TERMINAL_2_COL]
                    second_col_value = self.chart.loc[idx, TERMINAL_2_COL]

                    last_second, last_first = None, None

                    if back_first_col_value is None or back_first_col_value == ' ' or back_first_col_value == '':
                        if first_col_value is not None and first_col_value != ' ' and first_col_value != '':
                            index = idx - 1
                            while index > 0:
                                if back_first_col_value is not None and back_first_col_value != ' ' and back_first_col_value != '':
                                    break
                                else:
                                    index -= 1
                                back_first_col_value = self.chart.loc[index, TERMINAL_1_COL]
                            last_first = self.chart.loc[index, TERMINAL_1_COL]
                    if back_second_col_value is None or back_second_col_value == ' ' or back_second_col_value == '':
                        if second_col_value is not None and second_col_value != ' ' and second_col_value != '':
                            index = idx - 1
                            while index > 0:
                                if back_second_col_value is not None and back_second_col_value != ' ' and back_second_col_value != '':
                                    break
                                else:
                                    index -= 1
                                back_second_col_value = self.chart.loc[index, TERMINAL_2_COL]
                            last_second = self.chart.loc[index, TERMINAL_2_COL]

                    time_position = self.__time_changeover(idx, time, volume=quantity_harness,
                                                           last_first=last_first, last_second=last_second)

                if quantity is not None:
                    self.chart.loc[idx, AMOUNT_COL] = quantity_harness

                marking = self.chart.loc[idx, MARKING_COL]

                if type(marking) is str and 'б' in marking.lower():
                    if not no_more_white:
                        time_spend_white += time_position
                        if white_komax == 0:
                            pass
                            # raise Warning("No komaxes with white marking")
                        self.chart.loc[idx, KOMAX_COL] = white_komax
                    else:
                        # raise Warning("too many positions with white marking!")
                        return -1
                elif type(marking) is str and 'ч' in marking.lower():
                    if not no_more_black:
                        time_spend_black += time_position
                        if black_komax == 0:
                            # raise Warning("No komaxes with black marking")
                            pass
                        self.chart.loc[idx, KOMAX_COL] = black_komax
                    else:
                        # raise Warning("too many positions with black marking!")
                        return -1
                elif marking is None:
                    if black_komax_more:
                        if not no_more_black and black_komax != 0:
                            time_spend_black += time_position
                            self.chart.loc[idx, KOMAX_COL] = black_komax
                        elif not no_more_white and white_komax != 0:
                            time_spend_white += time_position
                            self.chart.loc[idx, KOMAX_COL] = white_komax
                        else:
                            # raise Warning("too many positions with no marking!")
                            return -1
                    else:
                        if not no_more_white and white_komax != 0:
                            time_spend_white += time_position
                            self.chart.loc[idx, KOMAX_COL] = white_komax
                        elif not no_more_black and black_komax != 0:
                            time_spend_black += time_position
                            self.chart.loc[idx, KOMAX_COL] = black_komax
                        else:
                            # raise Warning("too many positions with no marking!")
                            return -1

                last_row = idx == len(self.chart) - 1
                if time_spend_black > worker_time or last_row:
                    if black_komax == 0:
                        # raise Warning("No komaxes with black marking at all!")
                        return -1
                    alloc[black_komax][0] += time_spend_black
                    if not last_row:
                        black_komax_copy = black_komax
                        # find some free komax without pairing
                        for key, item in komaxes.items():
                            komax_work = komaxes[key][0] == 1
                            komax_black = komaxes[key][1] == 1
                            komax_pair = komaxes[key][2] == 1
                            current_komax = key == black_komax
                            less_worker_time = alloc[key][0] < worker_time
                            if komax_work and komax_black and not komax_pair and not current_komax and less_worker_time:
                                black_komax = key
                                time_spend_black = 0
                                break
                        # if not found, so find some pairing komax that is not fully loaded
                        if black_komax == black_komax_copy:
                            for key, item in komaxes.items():
                                komax_work = komaxes[key][0] == 1
                                komax_black = komaxes[key][1] == 1
                                komax_pair = komaxes[key][2] == 1
                                less_worker_time = komaxes[key][0] < worker_time
                                current_key = key == black_komax
                                if komax_work and komax_black and komax_pair and not current_key and less_worker_time:
                                    black_komax = key
                                    time_spend_black = 0
                                    break
                            if black_komax == black_komax_copy:
                                no_more_black = True

                if time_spend_white > worker_time or last_row:
                    if white_komax == 0:
                        # raise Warning("No komaxes with white marking at all")
                        return -1
                    alloc[white_komax][0] += time_spend_white
                    if not last_row:
                        white_komax_copy = white_komax
                        # find some free komax without pairing
                        for key, item in komaxes.items():
                            komax_work = komaxes[key][0] == 1
                            komax_black = komaxes[key][1] == 1
                            komax_pair = komaxes[key][2] == 1
                            current_komax = key == white_komax
                            less_worker_time = alloc[key][0] < worker_time
                            if komax_work and not komax_black and not komax_pair and not current_komax and less_worker_time:
                                white_komax = key
                                time_spend_white = 0
                                break
                        # if not found, so find some pairing komax that is not fully loaded
                        if white_komax == white_komax_copy:
                            for key, item in komaxes.items():
                                komax_work = komaxes[key][0] == 1
                                komax_black = komaxes[key][1] == 1
                                komax_pair = komaxes[key][2] == 1
                                less_worker_time = komaxes[key][0] < worker_time
                                current_key = key == white_komax
                                if komax_work and not komax_black and komax_pair and not current_key and less_worker_time:
                                    white_komax = key
                                    time_spend_white = 0
                                    break
                            if white_komax == white_komax_copy:
                                no_more_white = True

            else:
                pass

        return alloc

    def task_allocation_base(self, komaxes, quantity, time, hours):

        HARNESS_NUMBER_COL = "harness"
        AMOUNT_COL = "amount"
        KOMAX_COL = 'komax'
        MARKING_COL = "marking"
        TERMINAL_1_COL = 'wire_terminal_1'
        TERMINAL_2_COL = 'wire_terminal_2'


        """
        try:
            length2 = get_index(df, "Длина, мм (± 3мм) 2")[0]
        except TypeError:
            length2 = -1
        """

        # create dict of allocation
        alloc = {i: [0] for i in komaxes}
        worker_time = hours*60*60       # in seconds

        # get info of komaxes

        black_pairing_komax = 0
        white_pairing_komax = 0
        black_komax = 0
        white_komax = 0

        for key, item in komaxes.items():
            if komaxes[key][0] == 1:
                if komaxes[key][1] == 1:
                    if komaxes[key][2] == 1 and white_pairing_komax == 0:
                        white_pairing_komax = key
                    elif komaxes[key][2] == 0 and white_komax == 0:
                        white_komax = key
                elif komaxes[key][1] == 2:
                    if komaxes[key][2] == 1 and black_pairing_komax == 0:
                        black_pairing_komax = key
                    elif komaxes[key][2] == 0 and black_komax == 0:
                        black_komax = key

        if black_komax == 0:
            black_komax = black_pairing_komax
        if white_komax == 0:
            white_komax = white_pairing_komax

        amount_black_pairing_komax = 0
        amount_white_pairing_komax = 0

        amount_black_komax = 0
        amount_white_komax = 0

        for key, item in komaxes.items():
            if item[0] == 1 and item[1] == 1 and item[2] == 1:
                amount_white_pairing_komax += 1
            elif item[0] == 1 and item[1] == 2 and item[2] == 1:
                amount_black_pairing_komax += 1
            elif item[0] == 1 and item[1] == 1 and item[2] == 0:
                amount_white_komax += 1
            elif item[0] == 1 and item[1] == 2 and item[2] == 0:
                amount_black_komax += 1

        black_pairing_komax_more = amount_black_pairing_komax >= amount_white_pairing_komax
        black_komax_more = amount_black_komax >= amount_white_komax

        # adding columns with harnesses amount and komax number alloc


        self.chart[AMOUNT_COL] = ''

        self.chart[KOMAX_COL] = ''

        #count time for paired positions
        """
        if length2 == -1:
            pass
        else:
            time_spend_black_pairing = 0
            time_spend_white_pairing = 0

            no_more_black_pairing = False
            no_more_white_pairing = False

            for row in range(1, rows):
                value = df.iloc[row, length2]
                quantity_harness = 0
                time_position = 0
                if value is None:
                    pass
                else:
                    harness = df.ix[row, harness_number_col]
                    # quantity_harness = 1
                    try:
                        quantity_harness = quantity[harness]
                    except KeyError:
                        text_warn_rus_1 = "Проверь номера жгутов в файле кол-ва жгутов."
                        text_warn_rus_2 = "Не найден жгут номер {}".format(harness)
                        write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1 + text_warn_rus_2)
                        raise KeyError("Check if harness number is correct in \"Количество.xlsx\" file! : {}".format(harness))

                    time_position = time_changeover(df, row, column_numbers, time, volume=quantity_harness, pairing=True)

                    df.iloc[row, cols - 2] = quantity_harness

                    marking = df.iloc[row, mark_col]
                    if type(marking) is str and 'б' in marking.lower():
                        if not no_more_white_pairing:
                            time_spend_white_pairing += time_position
                            if white_pairing_komax == 0:
                                text_warn_rus_1 = "Нет комаксов с белой маркировкой и возможностью спарки!"
                                write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                                raise Warning("No komaxes with white marking and pairing ability")
                            df.iloc[row, cols - 1] = white_pairing_komax
                        else:
                            #TODO: replace "спарки" with "спаривание" everywhere please!
                            text_warn_rus_1 = "Слишком много позиций для спарки с белой маркировкой!"
                            write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                            # raise Warning("too many positions for pairing with white marking!")
                    elif type(marking) is str and 'ч' in marking.lower():
                        if not no_more_black_pairing:
                            time_spend_black_pairing += time_position
                            if black_pairing_komax == 0:
                                text_warn_rus_1 = "Нет комаксов с возможностью спарки и чёрной маркировкой!"
                                write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                                raise Warning("No komaxes with black marking and pairing ability")
                            df.iloc[row, cols - 1] = black_pairing_komax
                        else:
                            text_warn_rus_1 = "Слишком много позиций для спарки с чёрной маркировкой!"
                            write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                            # raise Warning("too many positions for pairing with black marking!")
                    elif marking is None:
                        # count whether more white or black paired komaxes
                        if black_pairing_komax_more:
                            if not no_more_black_pairing and black_pairing_komax != 0:
                                time_spend_black_pairing += time_position
                                df.iloc[row, cols - 1] = black_pairing_komax
                            elif not no_more_white_pairing and white_pairing_komax != 0:
                                time_spend_white_pairing += time_position
                                df.iloc[row, cols - 1] = white_pairing_komax
                            else:
                                text_warn_rus_1 = "Слишком много позиций для спарки без маркировки!"
                                text_warn_rus_2 = "Невозможно их распределить по komax"
                                write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1 + text_warn_rus_2)
                                # raise Warning("too many positions for pairing with no marking!")
                        else:
                            if not no_more_white_pairing and white_pairing_komax != 0:
                                time_spend_white_pairing += time_position
                                df.iloc[row, cols - 1] = white_pairing_komax
                            elif not no_more_black_pairing and black_pairing_komax != 0:
                                time_spend_black_pairing += time_position
                                df.iloc[row, cols - 1] = black_pairing_komax

                            else:
                                text_warn_rus_1 = "Слишком много позиций для спарки без маркировки!"
                                text_warn_rus_2 = "Невозможно их распределить по komax"
                                write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1 + text_warn_rus_2)
                                # raise Warning("too many positions for pairing with no marking!")

                    last_row = row == rows - 1
                    if time_spend_black_pairing > worker_time or last_row:
                        if black_pairing_komax == 0:
                            text_warn_rus_1 = "Нет komax с чёрной маркировкой и возможностью спарки!"
                            write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                            raise Warning("No komaxes with black marking and pairing ability at all!")
                        alloc[black_pairing_komax][0] += time_spend_black_pairing
                        if not last_row:
                            for key, item in komaxes.items():
                                komax_work = komaxes[key][0] == 1
                                komax_black = komaxes[key][1] == 2
                                komax_pair = komaxes[key][2] == 1
                                current_komax = key == black_pairing_komax
                                less_worker_time = alloc[key][0] < worker_time
                                if komax_work and komax_black and komax_pair and not current_komax and less_worker_time:
                                    black_pairing_komax = key
                                    time_spend_black_pairing = 0
                                    break
                                else:
                                    no_more_black_pairing = True

                    if time_spend_white_pairing > worker_time or last_row:
                        if white_pairing_komax == 0:
                            text_warn_rus_1 = "Нет комаксов с возможностью спаривания и белой маркировкой!"
                            write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1)
                            raise Warning("No komaxes with white marking and pairing ability at all!")
                        alloc[white_pairing_komax][0] += time_spend_white_pairing
                        if not last_row:
                            for key, item in komaxes.items():
                                komax_work = komaxes[key][0] == 1
                                komax_white = komaxes[key][1] == 1
                                komax_pair = komaxes[key][2] == 1
                                current_komax = key == white_pairing_komax
                                less_worker_time = alloc[key][0] < worker_time
                                if komax_work and komax_white and komax_pair and not current_komax and less_worker_time:
                                    white_pairing_komax = key
                                    time_spend_black_pairing = 0
                                    break
                                else:
                                    no_more_white_pairing = True
        """

        # count time for non-paired positions

        time_spend_black = 0
        time_spend_white = 0

        no_more_black = False
        no_more_white = False

        first_black = False
        first_white = False

        for idx, row in self.chart.iterrows():
            # value = df.iloc[row, length2]
            value = None
            if value is None:
                harness = self.chart.loc[idx, HARNESS_NUMBER_COL]

                if quantity is not None:
                    quantity_harness = quantity[harness]
                else:
                    quantity_harness = self.chart.loc[idx, AMOUNT_COL]
                """
                try:
                    quantity_harness = quantity[harness]
                except KeyError:
                    raise KeyError("Check if harness number is correct in \"quantity.xlsx\" file! : {}".format(harness))
                """

                if ('черный' in self.chart.loc[idx, MARKING_COL].lower() or 'чёрный' in self.chart.loc[idx, MARKING_COL].lower()) and not first_black:
                    first_black = True
                    time_position = self.__time_changeover(idx, time, volume=quantity_harness, full=True)
                elif 'белый' in self.chart.loc[idx, MARKING_COL].lower() and not first_white:
                    first_white = True
                    time_position = self.__time_changeover(idx, time, volume=quantity_harness, full=True)
                else:
                    #TODO: Try to make func for this part
                    back_first_col_value = self.chart.loc[idx - 1, TERMINAL_1_COL]
                    first_col_value = self.chart.loc[idx, TERMINAL_1_COL]

                    back_second_col_value = self.chart.loc[idx - 1, TERMINAL_2_COL]
                    second_col_value = self.chart.loc[idx, TERMINAL_2_COL]

                    last_second, last_first = None, None

                    if back_first_col_value is None or back_first_col_value == ' ' or back_first_col_value == '':
                        if first_col_value is not None and first_col_value != ' ' and first_col_value != '':
                            index = idx - 1
                            while index > 0:
                                if back_first_col_value is not None and back_first_col_value != ' ' and back_first_col_value != '':
                                    break
                                else:
                                    index -= 1
                                back_first_col_value = self.chart.loc[index, TERMINAL_1_COL]
                            last_first = self.chart.loc[index, TERMINAL_1_COL]
                    if back_second_col_value is None or back_second_col_value == ' ' or back_second_col_value == '':
                        if second_col_value is not None and second_col_value != ' ' and second_col_value != '':
                            index = idx - 1
                            while index > 0:
                                if back_second_col_value is not None and back_second_col_value != ' ' and back_second_col_value != '':
                                    break
                                else:
                                    index -= 1
                                back_second_col_value = self.chart.loc[index, TERMINAL_2_COL]
                            last_second = self.chart.loc[index, TERMINAL_2_COL]

                    time_position = self.__time_changeover(idx, time, volume=quantity_harness,
                                                           last_first=last_first, last_second=last_second)

                self.chart.loc[idx, AMOUNT_COL] = quantity_harness
                marking = self.chart.loc[idx, MARKING_COL]

                if type(marking) is str and 'б' in marking.lower():
                    if not no_more_white:
                        time_spend_white += time_position
                        if white_komax == 0:
                            pass
                            # raise Warning("No komaxes with white marking")
                        else:
                            self.chart.loc[idx, KOMAX_COL] = white_komax
                    else:
                        # raise Warning("too many positions with white marking!")
                        pass
                elif type(marking) is str and 'ч' in marking.lower():
                    if not no_more_black:
                        time_spend_black += time_position
                        if black_komax == 0:
                            # raise Warning("No komaxes with black marking")
                            pass
                        else:
                            self.chart.loc[idx, KOMAX_COL] = black_komax
                    else:
                        # raise Warning("too many positions with black marking!")
                        pass
                elif marking is None:
                    if black_komax_more:
                        if not no_more_black and black_komax != 0:
                            time_spend_black += time_position
                            self.chart.loc[idx, KOMAX_COL] = black_komax
                        elif not no_more_white and white_komax != 0:
                            time_spend_white += time_position
                            self.chart.loc[idx, KOMAX_COL] = white_komax
                        else:
                            # raise Warning("too many positions with no marking!")
                            pass
                    else:
                        if not no_more_white and white_komax != 0:
                            time_spend_white += time_position
                            self.chart.loc[idx, KOMAX_COL] = white_komax
                        elif not no_more_black and black_komax != 0:
                            time_spend_black += time_position
                            self.chart.loc[idx, KOMAX_COL] = black_komax
                        else:
                            # raise Warning("too many positions with no marking!")
                            pass

                last_row = idx == len(self.chart) - 1
                if time_spend_black > worker_time or last_row:
                    if black_komax == 0:
                        # raise Warning("No komaxes with black marking at all!")
                        pass
                    else:
                        alloc[black_komax][0] += time_spend_black
                    if not last_row:
                        black_komax_copy = black_komax
                        # find some free komax without pairing
                        for key, item in komaxes.items():
                            komax_work = komaxes[key][0] == 1
                            komax_black = komaxes[key][1] == 2
                            komax_pair = komaxes[key][2] == 1
                            current_komax = key == black_komax
                            less_worker_time = alloc[key][0] < worker_time
                            if komax_work and komax_black and not komax_pair and not current_komax and less_worker_time:
                                black_komax = key
                                time_spend_black = 0
                                break
                        # if not found, so find some pairing komax that is not fully loaded
                        if black_komax == black_komax_copy:
                            for key, item in komaxes.items():
                                komax_work = komaxes[key][0] == 1
                                komax_black = komaxes[key][1] == 2
                                komax_pair = komaxes[key][2] == 1
                                less_worker_time = komaxes[key][0] < worker_time
                                current_key = key == black_komax
                                if komax_work and komax_black and komax_pair and not current_key and less_worker_time:
                                    black_komax = key
                                    time_spend_black = 0
                                    break
                            if black_komax == black_komax_copy:
                                no_more_black = True

                if time_spend_white > worker_time or last_row:
                    if white_komax == 0:
                        # raise Warning("No komaxes with white marking at all")
                        pass
                    else:
                        alloc[white_komax][0] += time_spend_white
                    if not last_row:
                        white_komax_copy = white_komax
                        # find some free komax without pairing
                        for key, item in komaxes.items():
                            komax_work = komaxes[key][0] == 1
                            komax_black = komaxes[key][1] == 2
                            komax_pair = komaxes[key][2] == 1
                            current_komax = key == white_komax
                            less_worker_time = alloc[key][0] < worker_time
                            if komax_work and not komax_black and not komax_pair and not current_komax and less_worker_time:
                                white_komax = key
                                time_spend_white = 0
                                break
                        # if not found, so find some pairing komax that is not fully loaded
                        if white_komax == white_komax_copy:
                            for key, item in komaxes.items():
                                komax_work = komaxes[key][0] == 1
                                komax_black = komaxes[key][1] == 2
                                komax_pair = komaxes[key][2] == 1
                                less_worker_time = komaxes[key][0] < worker_time
                                current_key = key == white_komax
                                if komax_work and not komax_black and komax_pair and not current_key and less_worker_time:
                                    white_komax = key
                                    time_spend_white = 0
                                    break
                            if white_komax == white_komax_copy:
                                no_more_white = True

            else:
                pass

        return alloc

    def get_df_sorted_by_komax(df, alloc):
        try:
            komax_col = get_index(df, "komax")[0]
        except TypeError:
            komax_col = -1

        if komax_col == -1:
            return komax_col

        rows, cols = df.shape
        df_dict = {}

        for key, item in alloc.items():
            if item[0] > 0:
                df_dict[key] = pd.DataFrame(df[0:][:1])

        for row in range(1, rows):
            komax_number = df.iloc[row, komax_col]
            if komax_number in df_dict:
                temp_df = df[0:][row:row + 1]
                df_dict[komax_number] = df_dict[komax_number].merge(temp_df, how='outer')
            else:
                pass

        return df_dict

class KomaxLoading:
    __komaxes_dict = None
    __allocation_dict = None

    def __init__(self, komaxes):
        self.__komaxes_dict = komaxes
        self.__allocation_dict = {i: [0] for i in self.__komaxes_dict}

    def get_komaxes_dict(self):
        return self.__komaxes_dict

    def get_allocation_dict(self):
        return self.__allocation_dict

    def get_komaxes_allocation_by_komax_group_square(self, group_of_square):
        """
        Returns allocation dict by group of square

        :param group_of_square: int
        :return: alloc_dict
        """

        suitable_komaxes = list()
        group_of_square = str(group_of_square)

        for komax, params in self.__komaxes_dict.items():
            if group_of_square in params[3]:
                suitable_komaxes.append(komax)

        output_allocation = dict()
        output_komaxes = dict()

        for komax in suitable_komaxes:
            output_allocation[komax] = self.__allocation_dict[komax]
            output_komaxes[komax] = self.__komaxes_dict[komax]

        return (output_komaxes, output_allocation)

    def set_allocation(self, new_allocation):
        for komax, time in new_allocation.items():
            self.__allocation_dict[komax][0] = time[0]

class KomaxAppTaskName:
    task_name = None

    def __init__(self, task_name):
        self.task_name = task_name

    def check_task_name(self):
        if type(self.task_name) is int:
            return True
        elif type(self.task_name) is str:
            for symbol in self.task_name:
                if is_digit(symbol) or is_letter(symbol) or symbol == '-':
                    continue
                else:
                    return False

            return True
        else:
            return False

def is_digit(symbol: str):
    if symbol >= '0' and symbol <= '9':
        return True
    else:
        return False


def is_letter(symbol: str):
    symbol_lower = symbol.lower()
    if (symbol_lower >= 'a' and symbol_lower <= 'z') or (symbol_lower >= 'а' and symbol_lower <= 'я'):
        return True
    else:
        return False

def group_by(square):
    if square is not None:
        square = float(square)

    if square <= 1.0:
        return 1
    elif 1.0 < square <= 2.5:
        return 2
    elif square > 2.5:
        return 3

def empty(string):
    return False if string is not None and len(string) > 0 else True

def get_time_from(time_df):
    temp_dict = time_df.to_dict()

    output_dict = {}
    for idx, action in temp_dict['action'].items():
        output_dict[action] = temp_dict['time'][idx]

    return output_dict

def get_komaxes_from(komax_df):
    temp_dict = komax_df.to_dict()
    output_dict = {}

    for idx, komax_num in temp_dict['number'].items():
        status = (1 if temp_dict['status'][idx] == 'Work' else 0)
        marking = (1 if temp_dict['marking'][idx] == 'White' else 2)
        pairing = (1 if temp_dict['pairing'][idx] == 'Yes' else 0)


        output_dict[komax_num] = (status, marking, pairing)

    return output_dict

def get_amount_from(amount_df):
    temp_dict = amount_df.to_dict()
    output_dict = {}

    for idx, harness in temp_dict['harness'].items():
        output_dict[harness] = temp_dict['amount'][idx]

    return output_dict



