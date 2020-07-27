import openpyxl
import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from django.conf import settings
import time



class OutProcess:
    workbook = None
    worksheet = None
    dataframe = None

    def __init__(self, file=None):
        if file is not None and not isinstance(file, pd.DataFrame):
            self.workbook = load_workbook(file)
            self.worksheet = self.workbook.active
            self.dataframe = None
        elif file is not None and isinstance(file, pd.DataFrame):
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            self.dataframe = file
        else:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active

    def __style_xlsx(self, width=12, height=45, font_size=14, font_name='GOST type A', font_style='bold'):

        ws = self.worksheet

        # set the cell fonts and borders and alignment

        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        harness_number_col = 0
        armirovka1_col = 0
        armirovka2_col = 0
        terminal1_col = 0
        terminal2_col = 0
        color_col = 0
        komax_col = 0
        position_col = 0
        apl1_col = 0
        apl2_col = 0
        i = 0

        other_specials_width = width + 8
        harness_number_width = other_specials_width + 5

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border
                alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
                cell.alignment = alignment_obj

                value = cell.value
                if value == "Номер жгута":
                    harness_number_col = i
                elif value == "Наконечник 1":
                    terminal1_col = i
                elif value == "Наконечник 2":
                    terminal2_col = i
                elif value == "Уплотнитель 1":
                    armirovka1_col = i
                elif value == "Уплотнитель 2":
                    armirovka2_col = i
                elif value == "Аппликатор 1":
                    apl1_col = i
                elif value == "Аппликатор 2":
                    apl2_col = i
                elif value == "Цвет":
                    color_col = i
                elif value == "№ п/п":
                    position_col = i
                elif value == "komax":
                    komax_col = i
                elif value == "Вид провода":
                    cell.value = "Вид\nпровода"
                elif value == "Длина, мм (± 3мм)":
                    cell.value = "Длина, мм\n(± 3мм)"
                elif value == "Частичное снятие 1":
                    cell.value = "Частичное\nснятие 1"
                elif value == "Частичное снятие 2":
                    cell.value = "Частичное\nснятие 2"
                elif value == "№ провода":
                    cell.value = "№\nпровода"

                curr_col = cell.col_idx - 1
                bold = True if font_style == 'bold' else False
                if curr_col == terminal2_col or curr_col == terminal1_col or curr_col == armirovka2_col or curr_col == armirovka1_col:
                    cell.font = Font(size=font_size, bold=bold, name=font_name)
                else:
                    cell.font = Font(size=font_size, bold=bold, name=font_name)

                i += 1
            i = 0

        # set the dimension of cells by row and col

        row_count = ws.max_row
        column_count = ws.max_column
        columns_medium_width = [4, 10, 14]

        for row in range(1, row_count + 1):
            ws.row_dimensions[row].height = height
        for col in range(column_count):
            if col == harness_number_col:
                ws.column_dimensions[chr(ord('A') + col)].width = harness_number_width
            elif col == apl1_col or col == apl2_col:
                ws.column_dimensions[chr(ord('A') + col)].width = other_specials_width + 1
            elif col == armirovka1_col or col == armirovka2_col or col == terminal1_col or col == terminal2_col:
                ws.column_dimensions[chr(ord('A') + col)].width = other_specials_width
            elif col in columns_medium_width:
                ws.column_dimensions[chr(ord('A') + col)].width = other_specials_width - 4
            elif col == position_col or col == color_col or col == komax_col:
                ws.column_dimensions[chr(ord('A') + col)].width = width - 3
            else:
                ws.column_dimensions[chr(ord('A') + col)].width = width

    def __read_dataframe(self):
        if isinstance(self.dataframe, pd.DataFrame):
            if 'task' in self.dataframe:
                self.dataframe.drop(columns=['task'], inplace=True)
            if 'id' in self.dataframe:
                self.dataframe.drop(columns=['id'], inplace=True)

            for r in dataframe_to_rows(self.dataframe, index=True, header=True):
                self.worksheet.append(r)

    def __translate_cols_to_rus(self):
        first_row = self.worksheet[1]

        for cell in first_row:
            value = cell.value
            if value == 'harness':
                cell.value = "Номер жгута"
            elif value == 'marking':
                cell.value = "Маркировка"
            elif value == 'wire_type':
                cell.value = "Вид\nпровода"
            elif value == 'wire_terminal_1':
                cell.value = "Наконечник 1"
            elif value == 'wire_terminal_2':
                cell.value = "Наконечник 2"
            elif value == 'wire_seal_1':
                cell.value = "Уплотнитель 1"
            elif value == 'wire_seal_2':
                cell.value = "Уплотнитель 2"
            elif value == 'aplicator_1':
                cell.value = "Аппликатор 1"
            elif value == "aplicator_2":
                cell.value = "Аппликатор 2"
            elif value == 'wire_color':
                cell.value = "Цвет"
            elif value == 'wire_number':
                cell.value = "№ п/п"
            elif value == "wire_length":
                cell.value = "Длина, мм\n(± 3мм)"
            elif value == "wire_cut_length_1":
                cell.value = "Частичное\nснятие 1"
            elif value == "wire_cut_length_2":
                cell.value = "Частичное\nснятие 2"
            elif value == 'wire_square':
                cell.value = "Сечение"
            elif value == 'notes':
                cell.value = "Примечание"
            elif value == 'amount':
                cell.value = 'Количество'
            elif value == 'done':
                cell.value = 'Сделано'
            elif value == 'armirovka_1':
                cell.value = 'Армировка 1'
            elif value == 'armirovka_2':
                cell.value = 'Армировка 2'
            elif value == 'tube_len_1':
                cell.value = "Длина трубки, L (мм) 1"
            elif value == 'tube_len_2':
                cell.value = "Длина трубки, L (мм) 2"

    def get_task_xl(self):
        if self.dataframe is not None:
            self.__read_dataframe()
            self.__translate_cols_to_rus()

        self.__style_xlsx()

        self.worksheet.delete_rows(2, 1)
        return self.workbook

    def get_kappa_task_xl(self):
        if self.dataframe is not None:
            cols_to_drop = [
                'wire_seal_1',
                'wire_seal_2',
                'wire_terminal_1',
                'wire_terminal_2',
                'aplicator_1',
                'aplicator_2',
            ]
            self.dataframe = self.dataframe.drop(cols_to_drop, axis=1)
            self.__read_dataframe()
            self.__translate_cols_to_rus()

        self.__style_xlsx()
        self.worksheet.delete_rows(2, 1)
        return self.workbook


    def get_harness_chart_xl(self):
        if self.dataframe is not None:
            self.__read_dataframe()
            self.__translate_cols_to_rus()

        self.__style_xlsx()

        self.worksheet.delete_rows(2, 1)
        return self.workbook

    def __create_label(self, dict_values, start_row=0):
        ws = self.worksheet
        now = datetime.datetime.now()
        font_size = 9
        font_name = 'Calibri'

        rows = 11

        ws.column_dimensions[chr(ord('A') + 0)].width = 8.5703125
        ws.column_dimensions[chr(ord('A') + 1)].width = 4.140625
        ws.column_dimensions[chr(ord('A') + 2)].width = 8.5703125
        ws.column_dimensions[chr(ord('A') + 3)].width = 4.5703125

        for row in range(1, rows):
            ws.row_dimensions[start_row + row].height = 11.25
            if row == 1:
                ws.row_dimensions[start_row + row].height = 15.75
                ws['A{}'.format(start_row + row)].font = Font(size=font_size + 2, bold=False, name=font_name)
                ws['A{}'.format(start_row + row)].value = "PRETTL group"
                ws['C{}'.format(start_row + row)].font = Font(size=font_size + 2, bold=False, name=font_name)
                ws['C{}'.format(start_row + row)].value = 'Komax {}'.format(dict_values['komax'])
            elif row == 2:
                ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
                ws['A{}'.format(start_row + row)].value = "№ жгута"
                ws['B{}'.format(start_row + row)].font = Font(size=font_size + 2, bold=True, name=font_name)
                ws['B{}'.format(start_row + row)].value = dict_values['Номер жгута']
            elif row == 3:
                ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
                ws['A{}'.format(start_row + row)].value = "№ провода"
                # expected to add length to ticket, but need more info
                # ws['D{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
                # ws['D{}'.format(start_row + row)].value = 'L = {}'.format(dict_values['Длина'])
                ws['B{}'.format(start_row + row)].font = Font(size=font_size + 2, bold=True, name=font_name)
                ws['B{}'.format(start_row + row)].value = dict_values['№\nпровода']

            if row == 4:
                ws['A{}'.format(start_row + row)].value = "Сечение"
                ws['B{}'.format(start_row + row)].font = Font(size=font_size + 1, bold=True, name=font_name)
                ws['C{}'.format(start_row + row)].value = "Цвет"
                ws['D{}'.format(start_row + row)].font = Font(size=font_size + 1, bold=True, name=font_name)
                ws['B{}'.format(start_row + row)].value = dict_values["Сечение"]
                ws['D{}'.format(start_row + row)].value = dict_values["Цвет"]
            elif row == 5:
                ws['A{}'.format(start_row + row)].value = "Длина:"
                ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
                ws['B{}'.format(start_row + row)].value = dict_values["Длина"]
                ws['B{}'.format(start_row + row)].font = Font(size=font_size, bold=True, name=font_name)

            elif row == 6:
                ws['A{}'.format(start_row + row)].value = "Армировка 1"
                ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
                ws['C{}'.format(start_row + row)].value = "Армировка 2"
                ws['C{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
            elif row == 7:
                ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
                ws['A{}'.format(start_row + row)].value = dict_values["Уплотнитель 1"]
                ws['C{}'.format(start_row + row)].value = dict_values["Уплотнитель 2"]
                ws['C{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
            elif row == 8:
                ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
                ws['A{}'.format(start_row + row)].value = "Наконечник 1"
                ws['C{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
                ws['C{}'.format(start_row + row)].value = "Наконечник 2"
            elif row == 9:
                ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
                ws['A{}'.format(start_row + row)].value = dict_values["Наконечник 1"]
                ws['C{}'.format(start_row + row)].value = dict_values["Наконечник 2"]
                ws['C{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)

            elif row == 10:
                ws['A{}'.format(start_row + row)].value = "Количество: {}".format(dict_values["Количество"])
                ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
                day = now.day
                month = now.month
                year = now.year
                ws['C{}'.format(start_row + row)].value = "Дата: " + '{}.{}.{}'.format(day, month, year)
                ws['C{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)

    def get_labels(self):
        rows, cols = self.dataframe.shape
        start_row = 0
        for row in range(rows):
            dict_values = {}
            dict_values["komax"] = self.dataframe.loc[row, 'komax']
            dict_values['Номер жгута'] = self.dataframe.loc[row, 'harness']
            dict_values['№\nпровода'] = self.dataframe.loc[row, 'wire_number']
            dict_values['Уплотнитель 1'] = self.dataframe.loc[row, 'wire_seal_1']
            dict_values["Уплотнитель 2"] = self.dataframe.loc[row, 'wire_seal_2']
            dict_values["Наконечник 1"] = self.dataframe.loc[row, 'wire_terminal_1']
            dict_values["Наконечник 2"] = self.dataframe.loc[row, 'wire_terminal_2']
            dict_values["Количество"] = self.dataframe.loc[row, 'amount']
            dict_values["Сечение"] = self.dataframe.loc[row, 'wire_square']
            dict_values["Цвет"] = self.dataframe.loc[row, 'wire_color']
            dict_values['Длина'] = self.dataframe.loc[row, 'wire_length']

            self.__create_label(dict_values, start_row=start_row)
            start_row += 10

        self.worksheet.page_margins.left, self.worksheet.page_margins.right = 0.4, 0.4
        self.worksheet.page_margins.top, self.worksheet.page_margins.bottom = 0.5, 0.5
        self.worksheet.page_margins.header, self.worksheet.page_margins.footer = 1.3, 1.3
        self.worksheet.page_setup.scale = 160

        return self.workbook