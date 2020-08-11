#read, write files
# -*- coding: utf-8 -*-

import math
import pandas as pd
import re
from openpyxl import load_workbook

COLUMN_NAMES = ("Примечание", "№ п/п", "Маркировка", "Вид провода", "№ провода", "Сечение", "Цвет",
                "Длина, мм (± 3мм)", "Уплотнитель 1", "Длина трубки, L (мм) 1", "Длина трубки, L (мм) 2",
                "Частичное снятие 1", "Частичное снятие 2", "Наконечник 1", "Аппликатор 1", "Уплотнитель 2",
                "Наконечник 2", "Аппликатор 2", "Номер жгута", "Армировка 1 (Трубка ПВХ, Тр.Терм., изоляторы)",
                "Армировка 2 (Трубка ПВХ, Тр.Терм., изоляторы)")

class FileReader:
    CABEL = "Кабель"
    CABEL_COLUMN = "Примечание"

    numeric_columns = [
        "Сечение",
        "Частичное снятие 1",
        "Частичное снятие 2",
        "Длина трубки, L (мм) 1",
        "Длина трубки, L (мм) 2"
    ]

    file = 0
    workbook_file = None
    worksheet_file = None
    dataframe_file = None

    def __init__(self, file):
        self.file = file
        self.workbook_file = load_workbook(file)
        self.worksheet_file = self.workbook_file.active

    def __save_file(self):
        self.workbook_file.save(self.file)

    #TODO: check func, there is suspect it is not correct
    def __delete_gray_xlsx(self):
        ws = self.worksheet_file

        for row in ws.iter_rows():
            for cell in row:
                white_tint = 0.0

                color = cell.fill.start_color.index
                if cell.fill.start_color.tint != white_tint or (color != '00000000' and color != 'FFFFFFFF'and color != 0) or color == 8 or color == 22:
                    value = cell.value

                    if type(value) is float or type(value) is int:
                        if math.floor(value) > 16:
                            cell.value = None
                        else:
                            pass
                    elif type(value) is str:
                        cell.value = ''
                    else:
                        cell.value = None

    def __detect_cabels_xlsx(self):
        ws = self.worksheet_file

        delete_rows = False

        for merged_coordinates in ws.merged_cells.ranges:
            for merged_cell in ws[merged_coordinates.coord]:
                rows_list = []
                for cell in merged_cell:
                    temp_string = cell.value
                    if type(temp_string) is str and "кабель" in temp_string.lower():
                        rows_list = re.findall('\d+', merged_coordinates.coord)
                        delete_rows = True

                # write cabel in unmerged cells
                if delete_rows:
                    ws.unmerge_cells(merged_coordinates.coord)
                    for row in rows_list:
                        ws['A{}'.format(row)].value = "Кабель"
                    break

            delete_rows = False

    def __close_xlsx(self):
        self.workbook_file.save(self.file)
        self.worksheet_file = None

    def __get_first_row_to_read(self):
        for row in self.worksheet_file.iter_rows():
            for cell in row:
                value = cell.value
                if type(value) is str and value.lower() == "примечание":
                    return cell.row

    def __check_numeric_cols(self, numberic_cols):
        for col in numberic_cols:
            for idx, row in self.dataframe_file.iterrows():
                if type(self.dataframe_file.loc[idx, col]) is str:
                    return False

        return True

    def __delete_positions_without_length(self):
        # delete positions without length

        idx_to_drop = []

        for index, row in self.dataframe_file.iterrows():
            for column in self.numeric_columns:
                self.dataframe_file.loc[index, column] = float(self.dataframe_file.loc[index, column])
            value = self.dataframe_file.loc[index, "Длина, мм (± 3мм)"]
            if value is not None and value != '' and value == value:
                self.dataframe_file.loc[index, "Длина, мм (± 3мм)"] = int(
                    self.dataframe_file.loc[index, "Длина, мм (± 3мм)"])
            else:
                idx_to_drop.append(index)

        # delete positions without length

        self.dataframe_file.drop(index=idx_to_drop, inplace=True)

        # recover indecies
        self.dataframe_file.index = pd.Index(range(self.dataframe_file.shape[0]))

    def __process_file(self):
        """
        Delete or add permanent columns

        This func deletes:
        - Delete rows and cols if NaN
        - Check if all columns exist
        :return:
        """

        # delete if column not in columns_names
        for column_name, column_data in self.dataframe_file.iteritems():
            if column_name not in COLUMN_NAMES:
                """if type(column_name) is str and 'Unnamed' in column_name:
                    column_name = int(column_name[-1:])"""
                self.dataframe_file.drop(column_name, axis=1, inplace=True)

        self.dataframe_file.dropna(axis=0, how='all', inplace=True)

        # fulfill NaN with 0s for numeric columns

        self.dataframe_file.loc[:, self.numeric_columns] = self.dataframe_file.loc[:, self.numeric_columns].fillna(value=0)

        # fulfill NaN with empty string for other columns

        self.dataframe_file.fillna(value='', inplace=True)

        self.__delete_positions_without_length()

    def __read_xl(self, start_row=0):
        cnt = 0
        ws = self.worksheet_file

        for row in ws.iter_rows():
            cnt += 1
            if cnt == start_row:
                break
            else:
                for cell in row:
                    if cell.value != None:
                        cell.value = None

        temp_df = pd.DataFrame(ws.values)

        temp_df.dropna(inplace=True, how='all', axis=0)
        temp_df.dropna(inplace=True, how='all', axis=1)
        temp_df.index = pd.Index(range(temp_df.shape[0]))

        temp_df.rename(columns=temp_df.loc[0, :], inplace=True)

        temp_df.drop(index=[0, 1], inplace=True)
        temp_df.index = pd.Index(range(temp_df.shape[0]))

        return temp_df

    def read_file_chart(self):
        if self.file.name.endswith('xlsx'):

            # xlsx processing

            self.__delete_gray_xlsx()
            self.__detect_cabels_xlsx()
            row_start = self.__get_first_row_to_read()

            # dataframe processing

            self.dataframe_file = self.__read_xl(start_row=row_start)

            self.__close_xlsx()
            self.__process_file()




"""
def fix_xlsx(in_file):
    tmpfd, tmp = tempfile.mkstemp(dir=os.path.dirname(in_file))
    os.close(tmpfd)
    filename = '[Content_Types].xml'
    data = ''
    with zipfile.ZipFile(in_file, 'r') as zin:
        with zipfile.ZipFile(tmp, 'w') as zout:
            for item in zin.infolist():
                if item.filename != filename:
                    zout.writestr(item, zin.read(item.filename))
                else:
                    data = zin.read(filename).decode()
    os.remove(in_file)
    os.rename(tmp, in_file)
    data = data.replace('/xl/sharedStrings.xml', '/xl/SharedStrings.xml')
    with zipfile.ZipFile(in_file, mode='a', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(filename, data)

def xls_to_xlsx(filepath):
    # filename = ntpath.basename(filepath)

    pd.read_excel(filepath).to_excel(filepath[:-3] + 'xlsx', index=False)

    wb = load_workbook(filepath[:-3] + 'xlsx')
    ws = wb.active
    df = pd.DataFrame(ws.values)

    number_col, number_row = get_index(df, "Длина", "(± 3мм)")

    for index, row in df.iterrows():
        if index > number_row and df[number_col][index] is str:
            if df[number_col][index][0] == '\'':
                number = df[number_col][index][1:]
                df[number_col][index] = int(number)

    for r in dataframe_to_rows(df, index=True, header=True):
        ws.append(r)

    wb.close()

def list_of_merged_cells(filepath):

    a1, a2, h1, h2 = -1, -1, -1, -1

    wb = open_workbook(filename=filepath)
    ws = wb.sheet_by_index(0)

    num_rows = ws.nrows - 1
    num_cells = ws.ncols - 1
    curr_row = -1

    while curr_row < num_rows:
        curr_row += 1
        curr_cell = -1
        while curr_cell < num_cells:
            curr_cell += 1
            value = ws.cell_value(curr_row, curr_cell)
            print(value, type(value))
            if type(value) is str and "наконечник" in value.lower():
                if '1' in value:
                    h1 = curr_cell
                elif '2' in value:
                    h2 = curr_cell
            elif type(value) is str and "армировка" in value.lower():
                if '1' in value:
                    a1 = curr_cell
                elif '2' in value:
                    a2 = curr_cell
    if h1 == -1 or h2 == -1 or a1 == -1 or a2 == -1:
        #TODO : strange exception handle. Check if it needed and delete it in another way.
        write_to_xlsx("Ошибка", "Выход/", "Наконечник или Армировка не найдены!")
        raise Warning("Наконечник или Армировка not found!")

    a1_merged, a2_merged, h1_merged, h2_merged = [], [], [], []

    for merge in ws.merged_cells:
        r1, r2, c1, c2 = merge
        if c1 == a1 and r1 != r2 - 1:
            a1_merged.append(r1)
        elif c1 == a2 and r1 != r2 - 1:
            a2_merged.append(r1)
        elif c1 == h1 and r1 != r2 - 1:
            h1_merged.append(r1)
        elif c1 == h2 and r1 != r2 - 1:
            h2_merged.append(r1)

    output_list = [a1_merged, h1_merged, a2_merged, h2_merged]
    column_list = [a1, h1, a2, h2]

    len_output_list = len(output_list)
    for i in range(len_output_list):
        col = column_list[i]
        for j in range(len(output_list[i])):
            row = output_list[i][j]

            ws.write(label=ws.cell_value(row, col))

    return output_list

#TODO: what to do with ПВА cells?
def delete_gray_xls(workbook, worksheet, worksheet_to_write, rows, cols):
    #delete gray cells in xls files

    wb = workbook
    ws = worksheet
    ws2 = worksheet_to_write
    num_rows = rows - 1
    # ws.nrows - 1
    num_cells = cols - 1
    # ws.ncols - 1
    curr_row = -1

    col_restricted_to_delete_1 = -1
    col_restricted_to_delete_2 = -1

    while curr_row < num_rows:
        curr_row += 1
        curr_cell = -1
        while curr_cell < num_cells:
            curr_cell += 1
            xfx = ws.cell_xf_index(curr_row, curr_cell)
            xf = wb.xf_list[xfx]
            color = xf.background.pattern_colour_index
            value = ws.cell_value(curr_row, curr_cell)
            type_value = type(ws.cell_value(curr_row, curr_cell))

            if col_restricted_to_delete_1 == -1 and type_value is str and 'частичное снятие 1' in value.lower():
                col_restricted_to_delete_1 = curr_cell
            if col_restricted_to_delete_2 == -1 and type_value is str and 'частичное снятие 2' in value.lower():
                col_restricted_to_delete_2 = curr_cell

            if (color == 55 or color == 8) and curr_cell != col_restricted_to_delete_2 and curr_cell != col_restricted_to_delete_1:

                if (type_value is int or type_value is float) and math.floor(value) < 16:
                    pass
                else:
                    ws2.write(curr_row, curr_cell, None)

#TODO: what to do with ПВА cells?
def delete_gray_xlsx(worksheet):
    # deletes gray except ПВА

    ws = worksheet
    for row in ws.iter_rows():
        for cell in row:
            white_tint = 0.0

            color = cell.fill.start_color.index
            if cell.fill.start_color.tint != white_tint or color == 'FFA6A6A6' or color == 8 or color == 22:
                value = cell.value

                if type(value) is float or type(value) is int:
                    if math.floor(value) > 16:
                        cell.value = None
                else:
                    cell.value = None
            if cell.value == "ПВА":
                pass

# read file КРП (chart)
def read_file_chart(filepath):
    df1, df2 = 0, 0
    if filepath.endswith(".xls"):
        wb = open_workbook(filepath, formatting_info=True)
        ws = wb.sheet_by_index(0)
        wb_to_write = cp.copy(wb)
        ws_to_write = wb_to_write.get_sheet(0)
        rows, cols = ws.nrows, ws.ncols
        delete_gray_xls(wb, ws, ws_to_write, rows, cols)
        detect_cabel_in_xls(ws, ws_to_write)
        wb_to_write.save(filepath)
        try:
            xls_to_xlsx(filepath)
        finally:
            os.remove(filepath)
        try:
            filepath = filepath[:-3] + 'xlsx'
            df1 = pd.read_excel(filepath, header=None)
            df1.columns = range(df1.shape[1])

            skip_rows = get_index(df1, "Наконечник")[1]

            if skip_rows != -1:
                df2 = pd.read_excel(filepath, header=None, skiprows=skip_rows)
                cols = df2.shape[1]
                df2.columns = range(cols)
        finally:
            os.remove(filepath)
    elif filename.endswith(".xlsx"):
        wb = load_workbook(filepath)
        ws = wb.active
        delete_gray_xlsx(ws)
        detect_cabel_in_xlsx(ws)
        filename_len = len(filename)
        filepath = filepath[:-filename_len] + temp_filename
        filepath = filepath[:-3] + 'xlsx'
        wb.save(filepath)
        try:
            fix_xlsx(filepath)
            df1 = pd.read_excel(filepath, header=None)

            skip_rows = get_index(df1, "Наконечник")[1]
            if skip_rows != -1:
                df2 = pd.read_excel(filepath, header=None, skiprows=skip_rows)
                cols = df2.shape[1]
                df2.columns = range(cols)
        finally:
            os.remove(filepath)
    elif not (filename.endswith(".xlsx") or filename.endswith(".xls")):
        #TODO: replace "имеел" with "имеет" and add what file problem with
        write_to_xlsx("Ошибка", "Выход/", "Проверь, имеел ли файл расширение xlsx или xls!")
        raise TypeError("Check if file is xlsx or xls!")

    rows, cols = df1.shape
    harness_number = 0
    col_harness, row_harness = get_index(df1, "№ Жгута")

    field_name_passed = False
    for col in range(col_harness + 1, cols):
        value = df1.iloc[row_harness, col]
        unnamed = type(value) is str and "Unnamed" in value
        if value is not None and value != '' and not unnamed:
            harness_number = df1.iloc[row_harness][col]
            break

    column = df1.shape[1]


    while True:
        not_empty = False
        column_value = df1.iloc[:, column - 1]
        for value in column_value:
            if value is not None:
                not_empty = True
                break
        if not_empty:
            break
        else:
            column -= 1
            df1 = df1.drop(column, axis=1)

    # delete empty rows
    df2.dropna(axis=0, how='all')
    df2[column] = harness_number
    df2.iloc[0, column] = "Номер жгута"
    return df2

# read Нормы по времени (timing)
def read_file_time(filepath, operations_amount=11):
    try:
        df = pd.read_excel(filepath, header=None)
    except:
        #TODO: replace "Найстрока" with "Настройка"
        text_warning = "Can not read file with time in {}. Check if file exists and it's xls or xlsx".format(filepath)
        text_warn_rus_1 = "Не удаётся прочитать файл с трудоёмкостью {}.".format(filepath)
        text_warn_rus_2 = "Проверь, существует ли файл с трудоёмкостью и с расширением ли он xlsx"
        text_warn_rus_3 = "Проверь, существует ли папка Найстрока"
        write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1 + text_warn_rus_2 + text_warn_rus_3)
        raise Warning(text_warning)

    col_start, row_start = get_index(df, "Вид операции")
    row_start += 1

    time = dict()

    for row in range(row_start, row_start + operations_amount):
        operation_name = df.iloc[row, col_start]
        operation_time = df.iloc[row, col_start + 1]
        if "Нарезка" in operation_name:
            time["cut"] = operation_time  # нарезка и оштамповка провода на 1м
        elif "Смена направляющей" in operation_name:
            time["direction"] = operation_time  # смена управляющей
        elif "Смена аппликатора" in operation_name:
            time["aplicator"] = operation_time
        elif "Смена провода" in operation_name:
            time["wire"] = operation_time  # смена катушки с проводом
        elif "Обучение" in operation_name:
            time["learn"] = operation_time  # обучение комакса
        elif "Внесение параметров" in operation_name:
            time["set_up"] = operation_time  # внесение параметров провода
        elif "Смена катушки контактов" in operation_name:
            time["contact"] = operation_time  # смена катушки контактов
        elif "Смена уплотнительной" in operation_name:
            time["compact"] = operation_time  # смена уплотнительной втулки
        elif "Упаковка" in operation_name:
            time["pack"] = operation_time  # упаковка провода
        elif "Бирка" in operation_name:
            time["ticket"] = operation_time  # печать и установка(наклейка) бирки на провод
        elif "журнала" in operation_name:
            time["task"] = operation_time  # заполнение журнала с заданием

    return time

# read количество (quantity)
def read_file_quantity(filepath):
    try:
        df_quantity = pd.read_excel(filepath, header=None)
    except:
        #TODO: replace folder name "files" with "Настройка"
        text_warning = "Can not read file with quantity in {}. Check if file exists and it's xls or xlsx".format(filepath)
        text_warn_rus_1 = "Не удаётся прочитать файл с кол-вом жгутов {}.".format(filepath)
        text_warn_rus_2 = "Проверь, существует ли файл с кол-вом жгутов и с расширением ли он xlsx"
        text_warn_rus_3 = "Проверь, существует ли папка files"
        write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1 + text_warn_rus_2 + text_warn_rus_3)
        raise Warning(text_warning)

    output = dict()

    rows, cols = df_quantity.shape

    for row in range(1, rows):

        harness_number = df_quantity.iloc[row, 0]
        harness_number = str(harness_number)
        quantity = df_quantity.iloc[row, 1]

        output[harness_number] = quantity

    time_shift = df_quantity.iloc[1, 2]

    return output, time_shift

# read конфигурация комакса
def read_file_config(filepath):
    try:
        df_config = pd.read_excel(filepath, header=None)
    except:
        #TODO : replace "files" name of folder to "Настройка"
        text_warning_1 = "Can not read file with configuration of komaxes in " + filepath
        text_warning_2 = "Check if file exists and it's xls or xlsx"
        text_warn_rus_1 = "Не удаётся прочитать файл с конфигурацией komax {}.".format(filepath)
        text_warn_rus_2 = "Проверь, существует ли файл с конфигурацией komax и с расширением ли он xlsx"
        text_warn_rus_3 = "Проверь, существует ли папка files"
        write_to_xlsx("Ошибка", "Выход/", text_warn_rus_1 + text_warn_rus_2 + text_warn_rus_3)
        raise Warning(text_warning_1 + text_warning_2)

    #TODO: check if file fulfilled correctly

    config = dict()

    rows, cols = df_config.shape

    for row in range(1, rows):
        komax = df_config.iloc[row, 0]
        status = df_config.iloc[row, 1]
        mark = df_config.iloc[row, 2]
        pair = df_config.iloc[row, 3]

        config[komax] = [status, mark, pair]

    return config

def write_to_xlsx(name='default_name', directory='', *args):
    i = 1

    if os.path.exists(directory):
        pass
    else:
        os.makedirs(directory)
    wb_output = Workbook()
    ws_output = wb_output.active
    for df in args:
        filepath = directory + name + '.xlsx'
        if os.path.exists(filepath):
            os.remove(filepath)


        if type(df) is pd.DataFrame or type(df) is pd.Series:
            for r in dataframe_to_rows(df, index=True, header=True):
                ws_output.append(r)
        elif type(df) is not None:
            name_of_cell = 'A{}'.format(i)
            ws_output[name_of_cell] = df
            i += 1

    wb_output.close()
    wb_output.save(filepath)

def write_to_pickle(name, directory='', *args):
    filepath = directory + name + '.xz'
    if os.path.exists(filepath):
        os.remove(filepath)
    for df in args:
        df.to_pickle(filepath, compression='xz')

def create_folder(current_directory, name="Default_name"):
    if type(name) is str:
        pass
    else:
        try:
            name = str(name)
        except:
            raise Warning("Shit with name!")

    filepath = current_directory + '/' + name

    if os.path.isdir(filepath):
        for filename in os.listdir(filepath):
            filepath_tmp = os.path.join(filepath, filename)
            os.remove(filepath_tmp)
        os.rmdir(filepath)

    os.mkdir(filepath)

def style_xlsx(name, directory='', width=12, height=45, font_size=14, font_name='GOST type A', font_style='bold'):

    wb = load_workbook(directory + name + '.xlsx')
    ws = wb.active
    ws.delete_rows(1, 2)
    ws.delete_cols(1)

    # set the cell fonts and borders and alignment

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    harness_number_col = 0
    armirovka1_col = 0
    armirovka2_col = 0
    terminal1_col = 0
    terminal2_col = 0
    color_col = 0
    komax_col = 0
    position_col = 0
    apl1_col = 0
    apl2_col = 0
    i = 0

    other_specials_width = width + 8
    harness_number_width = other_specials_width + 5

    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
            cell.alignment = alignment_obj

            value = cell.value
            if value == "Номер жгута":
                harness_number_col = i
            elif value == "Наконечник 1":
                terminal1_col = i
            elif value == "Наконечник 2":
                terminal2_col = i
            elif value == "Уплотнитель 1":
                armirovka1_col = i
            elif value == "Уплотнитель 2":
                armirovka2_col = i
            elif value == "Аппликатор 1":
                apl1_col = i
            elif value == "Аппликатор 2":
                apl2_col = i
            elif value == "Цвет":
                color_col = i
            elif value == "№ п/п":
                position_col = i
            elif value == "komax":
                komax_col = i
            elif value == "Вид провода":
                cell.value = "Вид\nпровода"
            elif value == "Длина, мм (± 3мм)":
                cell.value = "Длина, мм\n(± 3мм)"
            elif value == "Частичное снятие 1":
                cell.value = "Частичное\nснятие 1"
            elif value == "Частичное снятие 2":
                cell.value = "Частичное\nснятие 2"
            elif value == "№ провода":
                cell.value = "№\nпровода"


            curr_col = cell.col_idx - 1
            bold = True if font_style == 'bold' else False
            if curr_col == terminal2_col or curr_col == terminal1_col or curr_col == armirovka2_col or curr_col == armirovka1_col:
                cell.font = Font(size=font_size, bold=bold, name=font_name)
            else:
                cell.font = Font(size=font_size, bold=bold, name=font_name)

            i += 1
        i = 0

    # set the dimension of cells by row and col

    row_count = ws.max_row
    column_count = ws.max_column
    columns_medium_width = [0, 1, 3, 8, 10, 12, 14, 16]

    for row in range(1, row_count + 1):
        ws.row_dimensions[row].height = height
    for col in range(column_count):
        if col == harness_number_col:
            ws.column_dimensions[chr(ord('A') + col)].width = harness_number_width
        elif col == apl1_col or col == apl2_col:
            ws.column_dimensions[chr(ord('A') + col)].width = other_specials_width + 1
        elif col == armirovka1_col or col == armirovka2_col or col == terminal1_col or col == terminal2_col:
            ws.column_dimensions[chr(ord('A') + col)].width = other_specials_width
        elif col in columns_medium_width:
            ws.column_dimensions[chr(ord('A') + col)].width = other_specials_width - 4
        elif col == position_col or col == color_col or col == komax_col:
            ws.column_dimensions[chr(ord('A') + col)].width = width - 3
        else:
            ws.column_dimensions[chr(ord('A') + col)].width = width


    wb.close()
    wb.save(directory + name + '.xlsx')

def read_file_komax_terminals(filepath):
    try:
        df = pd.read_excel(filepath, skiprows=2)
    except:
        #TODO: add "или некорректного расширения"
        write_to_xlsx('Ошибка', 'Выход/', 'Файл с наконечниками на komax не найден!')
        raise Warning('File with komax terminals not found')

    df = df.replace({pd.np.nan: None})

    col = 1

    rows, cols = df.shape

    terminals_list = []

    for row in range(1, rows):
        value = df.iloc[row, col]
        if value is not None:
            terminals_list.append(value)

    return terminals_list

def create_label(ws, dict_values, start_row=0):
    now = datetime.datetime.now()
    font_size = 9
    font_name = 'Calibri'

    rows = 10

    ws.column_dimensions[chr(ord('A') + 0)].width = 8.5703125
    ws.column_dimensions[chr(ord('A') + 1)].width = 4.140625
    ws.column_dimensions[chr(ord('A') + 2)].width = 8.5703125
    ws.column_dimensions[chr(ord('A') + 3)].width = 4.5703125

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    for row in ws.iter_rows():
        for cell in row:
            cell.border = thin_border
            alignment_obj = cell.alignment.copy(horizontal='center', vertical='center')
            cell.alignment = alignment_obj

    for row in range(1, rows):
        ws.row_dimensions[start_row + row].height = 11.25
        if row == 1:
            ws.row_dimensions[start_row + row].height = 15.75
            ws.merge_cells('A{}:B{}'.format(start_row + row, start_row + row))
            ws.merge_cells('C{}:D{}'.format(start_row + row, start_row + row))
            ws['A{}'.format(start_row + row)].font = Font(size=font_size + 2, bold=False, name=font_name)
            ws['A{}'.format(start_row + row)].value = "PRETTL group"
            ws['C{}'.format(start_row + row)].font = Font(size=font_size + 2, bold=False, name=font_name)
            ws['C{}'.format(start_row + row)].value = 'Komax {}'.format(dict_values['komax'])
        elif row == 2:
            ws.merge_cells('B{}:D{}'.format(start_row + row, start_row + row))
            ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
            ws['A{}'.format(start_row + row)].value = "№ жгута"
            ws['B{}'.format(start_row + row)].font = Font(size=font_size + 2, bold=True, name=font_name)
            ws['B{}'.format(start_row + row)].value = dict_values['Номер жгута']
        elif row == 3:
            ws.merge_cells('B{}:D{}'.format(start_row + row, start_row + row))
            ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
            ws['A{}'.format(start_row + row)].value = "№ провода"
            ws['B{}'.format(start_row + row)].font = Font(size=font_size + 2, bold=True, name=font_name)
            ws['B{}'.format(start_row + row)].value = dict_values['№\nпровода']
        elif row == 5 or row == 6 or row == 7 or row == 8:
            ws.merge_cells('A{}:B{}'.format(start_row + row, start_row + row))
            ws.merge_cells('C{}:D{}'.format(start_row + row, start_row + row))

        if row == 4:
            ws['A{}'.format(start_row + row)].value = "Сечение"
            ws['B{}'.format(start_row + row)].font = Font(size=font_size + 1, bold=True, name=font_name)
            ws['C{}'.format(start_row + row)].value = "Цвет"
            ws['D{}'.format(start_row + row)].font = Font(size=font_size + 1, bold=True, name=font_name)
            ws['B{}'.format(start_row + row)].value = dict_values["Сечение"]
            ws['D{}'.format(start_row + row)].value = dict_values["Цвет"]
        elif row == 5:
            ws['A{}'.format(start_row + row)].value = "Армировка 1"
            ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
            ws['C{}'.format(start_row + row)].value = "Армировка 2"
            ws['C{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
        elif row == 6:
            ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
            ws['A{}'.format(start_row + row)].value = dict_values["Уплотнитель 1"]
            ws['C{}'.format(start_row + row)].value = dict_values["Уплотнитель 2"]
            ws['C{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
        elif row == 7:
            ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
            ws['A{}'.format(start_row + row)].value = "Наконечник 1"
            ws['C{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
            ws['C{}'.format(start_row + row)].value = "Наконечник 2"
        elif row == 8:
            ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
            ws['A{}'.format(start_row + row)].value = dict_values["Наконечник 1"]
            ws['C{}'.format(start_row + row)].value = dict_values["Наконечник 2"]
            ws['C{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)

        elif row == 9:
            ws.merge_cells('C{}:D{}'.format(start_row + row, start_row + row))
            ws.merge_cells('A{}:B{}'.format(start_row + row, start_row + row))
            ws['A{}'.format(start_row + row)].value = "Количество: {}".format(dict_values["Количество"])
            ws['A{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)
            day = now.day
            month = now.month
            year = now.year
            ws['C{}'.format(start_row + row)].value = "Дата: " + '{}.{}.{}'.format(day, month, year)
            ws['C{}'.format(start_row + row)].font = Font(size=font_size, bold=False, name=font_name)

def create_labels(df, name, directory=''):
    wb = Workbook()
    ws = wb.active

    col_dict = dict_column_numbers(df)
    komax_col = get_index(df, "komax")[0]
    amount_col = get_index(df, "Количество")[0]
    harness_col = get_index(df, "Номер жгута")[0]
    wire_num_col = get_index(df, "№ провода")[0]

    col_dict["Количество"] = amount_col
    col_dict["komax"] = komax_col
    col_dict['Номер жгута'] = harness_col
    col_dict['№\nпровода'] = wire_num_col

    rows, cols = df.shape
    start_row = 0
    for row in range(1, rows):
        dict_values = {}
        dict_values["komax"] = df.iloc[row, col_dict['komax']]
        dict_values['Номер жгута'] = df.iloc[row, col_dict['Номер жгута']]
        dict_values['№\nпровода'] = df.iloc[row, col_dict['№\nпровода']]
        dict_values['Уплотнитель 1'] = df.iloc[row, col_dict['Уплотнитель 1']]
        dict_values["Уплотнитель 2"] = df.iloc[row, col_dict["Уплотнитель 2"]]
        dict_values["Наконечник 1"] = df.iloc[row, col_dict["Наконечник 1"]]
        dict_values["Наконечник 2"] = df.iloc[row, col_dict["Наконечник 2"]]
        dict_values["Количество"] = df.iloc[row, col_dict["Количество"]]
        dict_values["Сечение"] = df.iloc[row, col_dict["Сечение"]]
        dict_values["Цвет"] = df.iloc[row, col_dict["Цвет"]]

        create_label(ws, dict_values, start_row=start_row)
        start_row += 9

    ws.page_margins.left, ws.page_margins.right = 0.4, 0.4
    ws.page_margins.top, ws.page_margins.bottom = 0.5, 0.5
    ws.page_margins.header, ws.page_margins.footer = 1.3, 1.3
    ws.page_setup.scale = 160
    wb.close()
    wb.save(directory + name + '.xlsx')

def detect_cabel_in_xlsx(worksheet):
    ws = worksheet

    delete_rows = False

    for merged_coordinates in ws.merged_cells.ranges:
        for merged_cell in ws[merged_coordinates.coord]:
            rows_list = []
            for cell in merged_cell:
                temp_string = cell.value
                if type(temp_string) is str and "кабель" in temp_string.lower():
                    rows_list = re.findall('\d+', merged_coordinates.coord)
                    delete_rows = True

            # write cabel in unmerged cells
            if delete_rows:
                ws.unmerge_cells(merged_coordinates.coord)
                for row in rows_list:
                    ws['A{}'.format(row)].value = "Кабель"
                break

        delete_rows = False

def detect_cabel_in_xls(worksheet, worksheet_to_write):
    ws = worksheet
    ws2 = worksheet_to_write

    for crange in ws.merged_cells:
        rlo, rhi, clo, chi = crange
        temp_string = ws.cell_value(rlo, clo)
        if type(temp_string) is str and "кабель" in temp_string.lower():
            for row in range(rlo, rhi):
                for col in range(clo, chi):
                    ws2.write(row, col, "Кабель")
                    
"""
